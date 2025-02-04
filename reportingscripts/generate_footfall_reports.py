import json
import os
import sys
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# Writing to an excel  
# sheet using Python 
from openpyxl import load_workbook
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
PC_REQUEST_PATH = "./requests/pc_request.json"
SH_REQUEST_PATH = "./requests/systemHealth.json"
MONTH_REQUEST_PATH = "./requests/month_event_request.json"
date=None
priorPeriodReportEnable=False
numberOfDayDelta=0
url = "https://localhost:9200/filebeat-*/_search/?size="
#url = "http://smargtech.in:36523/filebeat-*/_search/?size="
#url = "http://smargtechnologies.in:36523/filebeat-*/_search/?size="
headers = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzp1SDFWQlgtLUx6UnlTTTVCaGd5QQ=='}
def create_request_for_all_data(site_id, zone_ids, from_date, to_date):
    #json=json.loads("{\"query\":{\"bool\":{\"must\":[],\"filter\": [{\"term\": {\"SiteId\": 111}},{\"term\": {\"ZoneId\": 1}},{\"term"\: {\"isViolation\": 0}},{\"term\": {\"Direction.keyword\": \"IN\"}},{\"term\": {\"NotificationType\": \"7\"}},{\"range\": {\"DateTime\": {\"gte\": \"2020-09-26T00:00:00Z\",\"lte\": \"2020-12-26T11:00:00Z\",}}},],}},\"sort\": [{\"@timestamp\": {\"order\": \"desc\"}}],}")
    json_obj=None
    with open(PC_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][0]["term"]["SiteId"]=site_id
    json_obj["query"]["bool"]["filter"][1]["terms"]["CameraID"]=zone_ids
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    #print(json.dumps(json_obj))
    return json_obj
def create_request_for_systemHealth(site_id, from_date, to_date):
    #json=json.loads("{\"query\":{\"bool\":{\"must\":[],\"filter\": [{\"term\": {\"SiteId\": 111}},{\"term\": {\"ZoneId\": 1}},{\"term"\: {\"isViolation\": 0}},{\"term\": {\"Direction.keyword\": \"IN\"}},{\"term\": {\"NotificationType\": \"7\"}},{\"range\": {\"DateTime\": {\"gte\": \"2020-09-26T00:00:00Z\",\"lte\": \"2020-12-26T11:00:00Z\",}}},],}},\"sort\": [{\"@timestamp\": {\"order\": \"desc\"}}],}")
    json_obj=None
    with open(SH_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][1]["term"]["SiteId"]=site_id
    json_obj["query"]["bool"]["filter"][3]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][3]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    # print(str(json_obj))
    return json_obj
def mothly_weeklyData(site_id, f_date, t_date):
    w_row=2
    # weekCount = report_wb['WeekCount']
    weekCount =None
    if priorPeriodReportEnable:
        weekCount = report_wb['PriorCount']
    else:
        weekCount = report_wb['WeekCount']
    fromTimeOfSite=f_date.time()
    toTimeOfSite=t_date.time()
    fromDateOfSite=f_date.date()
    toDateOfSite=t_date.date()
    dayTime=toTimeOfSite.hour-fromTimeOfSite.hour
    isSameDay=False
    if(dayTime>0):
        isSameDay=True
    delta = timedelta(days=1)
    DLC=''
    DLW=''
    DLM=''
    while fromDateOfSite <= toDateOfSite:
        if not isSameDay and fromDateOfSite==toDateOfSite:
            fromDateOfSite +=delta
        else:
            dateWiseFromDate=str(fromDateOfSite)+"T"+str(fromTimeOfSite)
            dateWiseToDate=""
            if isSameDay:
                dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
                fromDateOfSite +=delta
            else:
                fromDateOfSite +=delta
                dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
            #Current Date Count
            dateWiseFromDate =datetime.strptime(dateWiseFromDate, "%Y-%m-%dT%H:%M:%S")
            dateWiseToDate =datetime.strptime(dateWiseToDate, "%Y-%m-%dT%H:%M:%S")
            current_payload = create_request_for_all_data(site_id, zonesId, dateWiseFromDate, dateWiseToDate)
            #print("--current_payload",current_payload)
            #print(url)
            current_response = requests.post(url+"0",json=current_payload,headers=headers,timeout=(300,300),verify=False)
            #print(current_response.json())
            current_response=current_response.json()
            current_total=current_response["aggregations"]["Total_Count"]["value"]
            
            if priorPeriodReportEnable:
                #Past Week Date Count
                datePastWeekFromDate = dateWiseFromDate - timedelta(days=numberOfDayDelta)
                datePastWeekToDate = dateWiseToDate - timedelta(days=numberOfDayDelta)
                # print(datePastWeekFromDate)
                # print(datePastWeekToDate)
                past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
                past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300),verify=False)
                past_week_response=past_week_response.json()
                past_week_total=past_week_response["aggregations"]["Total_Count"]["value"]
                
                new_f_date = datetime.strftime(dateWiseFromDate, "%d-%m-%Y")
                weekCount.cell(w_row+1, 1, w_row-1)
                weekCount.cell(w_row+1, 2, new_f_date)
                weekCount.cell(w_row+1, 3, current_total)
                weekCount.cell(w_row+1, 4, past_week_total)
                if w_row==2:
                    DLC=dateWiseFromDate.strftime('%Y-%m-%d')
                    DLW=datePastWeekFromDate.strftime('%Y-%m-%d')
                if True:
                    DLC_1="Selected ( "+str(DLC)+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
                    DLW_1="Prior Period ( "+str(DLW)+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
                    weekCount.cell(1, 3,DLC_1 )
                    weekCount.cell(1, 4,DLW_1 )
            else:
                #Past Week Date Count
                datePastWeekFromDate = dateWiseFromDate - timedelta(days=7)
                datePastWeekToDate = dateWiseToDate - timedelta(days=7)
                past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
                past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300),verify=False)
                past_week_response=past_week_response.json()
                past_week_total=past_week_response["aggregations"]["Total_Count"]["value"]
                
                #Past Month Count
                # datePastMonthFromDate = dateWiseFromDate + relativedelta(months=-1)
                # datePastMonthToDate = dateWiseToDate + relativedelta(months=-1)
                datePastMonthFromDate = dateWiseFromDate - timedelta(days=28)
                datePastMonthToDate = dateWiseToDate - timedelta(days=28)
                past_month_payload = create_request_for_all_data(site_id, zonesId, datePastMonthFromDate, datePastMonthToDate)
                past_month_response = requests.get(url+"0",json=past_month_payload,headers=headers,timeout=(300,300),verify=False)
                past_month_response=past_month_response.json()
                past_month_total=past_month_response["aggregations"]["Total_Count"]["value"]
                
                new_f_date = datetime.strftime(dateWiseFromDate, "%d-%m-%Y")
                weekCount.cell(w_row+1, 1, w_row-1)
                weekCount.cell(w_row+1, 2, dateWiseFromDate.strftime('%A'))
                weekCount.cell(w_row+1, 3, current_total)
                weekCount.cell(w_row+1, 4, past_week_total)
                weekCount.cell(w_row+1, 5, past_month_total)
                if w_row==2:
                    DLC=dateWiseFromDate.strftime('%Y-%m-%d')
                    DLW=datePastWeekFromDate.strftime('%Y-%m-%d')
                    DLM=datePastMonthFromDate.strftime('%Y-%m-%d')
                    
                if True:
                    DLC_1="Selected ( "+str(DLC)+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
                    DLW_1="Last Week ( "+str(DLW)+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
                    DLM_1="Last Month ( "+str(DLM)+" TO "+str(datePastMonthToDate.strftime('%Y-%m-%d'))+" )"
                    weekCount.cell(1, 3,DLC_1 )
                    weekCount.cell(1, 4,DLW_1 )
                    weekCount.cell(1, 5,DLM_1 )
            
            # allDataCountDict[new_f_date] = allCountsDict 
        w_row=w_row+1
     
def get_percentage_diff(previous, current):
    try:
        percentage = (previous-current)/current * 100
    except ZeroDivisionError:
        percentage = float(100)    
    return round(percentage,2)
def zone_wise_details(site_id, f_date, t_date):
    # t3 = t_date - f_date
    # t3 = t3.days
    datePastWeekFromDate=''
    datePastWeekToDate=''
    datePastYearFromDate=''
    datePastYearToDate=''
    fromTimeOfSite=f_date.time()
    toTimeOfSite=t_date.time()
    fromDateOfSite=f_date.date()
    toDateOfSite=t_date.date()
    dateWiseFromDate=str(fromDateOfSite)+"T"+str(fromTimeOfSite)
    dateWiseToDate=str(toDateOfSite)+"T"+str(toTimeOfSite)       
    #Current Date Count
    dateWiseFromDate =datetime.strptime(dateWiseFromDate, "%Y-%m-%dT%H:%M:%S")
    dateWiseToDate =datetime.strptime(dateWiseToDate, "%Y-%m-%dT%H:%M:%S")
    current_payload = create_request_for_all_data(site_id, zonesId, dateWiseFromDate, dateWiseToDate)
    current_response = requests.get(url+"0",json=current_payload,headers=headers,timeout=(300,300),verify=False)
    current_response=current_response.json()

    #print(json.dumps(current_payload))


    t_res=current_response["hits"]["total"]["value"]
    if t_res is None or t_res == 0:
        return None
    allZoneCountCDate=current_response["aggregations"]["Total_Count"]["value"]
    allZoneCountCurrentDate=current_response["aggregations"]["Zone_wise_total"]["buckets"]
    #print(allZoneCountCurrentDate)
    if priorPeriodReportEnable:
        #Past Week Date Count
        datePastWeekFromDate = dateWiseFromDate - timedelta(days=numberOfDayDelta)
        datePastWeekToDate = dateWiseToDate - timedelta(days=numberOfDayDelta)
        past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
        past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300),verify=False)
        past_week_response=past_week_response.json()
        allZoneCountPastWeek=past_week_response["aggregations"]["Zone_wise_total"]["buckets"]
        allZonePastWCount=past_week_response["aggregations"]["Total_Count"]["value"]
    else:
        #Past Week Date Count
        datePastWeekFromDate = dateWiseFromDate - timedelta(days=7)
        datePastWeekToDate = dateWiseToDate - timedelta(days=7)
        past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
        past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300),verify=False)
        past_week_response=past_week_response.json()
        allZoneCountPastWeek=past_week_response["aggregations"]["Zone_wise_total"]["buckets"]
        allZonePastWCount=past_week_response["aggregations"]["Total_Count"]["value"]
        #Past Year Count
        datePastYearFromDate = dateWiseFromDate - relativedelta(years=1)
        datePastYearToDate = dateWiseToDate - relativedelta(years=1)
        past_year_payload = create_request_for_all_data(site_id, zonesId, datePastYearFromDate, datePastYearToDate)
        past_year_response = requests.get(url+"0",json=past_year_payload,headers=headers,timeout=(300,300),verify=False)
        past_year_response=past_year_response.json()
        allZoneCountPastYear=past_year_response["aggregations"]["Zone_wise_total"]["buckets"]
        allZonePastYCount=past_year_response["aggregations"]["Total_Count"]["value"]
    #zone dashboard data
    zoned_row=75
    footfall_dash.cell(zoned_row+1, 3, allZoneCountCDate)
    diff_percent_prior_week = get_percentage_diff(allZoneCountCDate,allZonePastWCount)
    footfall_dash.cell(zoned_row+3, 3, str(diff_percent_prior_week)+"%")
    
    if not priorPeriodReportEnable:
        diff_percent_prior_year = get_percentage_diff(allZoneCountCDate,allZonePastYCount)
        footfall_dash.cell(zoned_row+5, 3, str(diff_percent_prior_year)+"%")        
    
    
    zone_wrow=1
    zone_list = {}
    zone_count_list={}
    maxzone_week = 0
    maxzone_year = 0
    zoneWiseCount = report_wb['Zone_Wise_Count']
    for zoneCountData in allZoneCountCurrentDate:
        zoneWiseCount.cell(zone_wrow+1, 1, zoneCountData["key"])
        zoneWiseCount.cell(zone_wrow+1, 2, zoneCountData["doc_count"])
        zoneWiseCount.cell(zone_wrow+1, 3, 1)
        zoneWiseCount.cell(zone_wrow+1, 5, 1)
        zone_list.update({zoneCountData["key"]:zone_wrow+1})
        zone_count_list.update({zoneCountData["key"]:zoneCountData["doc_count"]})
        zone_wrow = zone_wrow+1
    zone_listv = list(zone_count_list.values())
    zone_listk = list(zone_count_list.keys())
    maxzone_current = zone_listk[zone_listv.index(max(zone_listv))]
    for zoneCountDataWeek in allZoneCountPastWeek:
        if zoneCountDataWeek["key"] in zone_list.keys():
            zoneWiseCount.cell(zone_list[zoneCountDataWeek["key"]], 4, zoneCountDataWeek["doc_count"])
            zoneWiseCountWeek = zoneCountDataWeek["doc_count"]
            change_percent_week = get_percentage_diff(zone_count_list[zoneCountDataWeek["key"]],zoneWiseCountWeek)
            zoneWiseCount.cell(zone_list[zoneCountDataWeek["key"]], 3, change_percent_week/100)
            
            if zoneCountDataWeek["key"] == maxzone_current:
                maxzone_week = zoneCountDataWeek["doc_count"]    
        else:
            zone_list.update({zoneCountDataWeek["key"]:zone_wrow+1})
            zoneWiseCount.cell(zone_wrow+1, 1, zoneCountDataWeek["key"])
            zoneWiseCount.cell(zone_wrow+1, 4, zoneCountDataWeek["doc_count"])
            zone_count_list.update({zoneCountDataWeek["key"]:zoneCountDataWeek["doc_count"]})
            zoneWiseCount.cell(zone_list[zoneCountDataWeek["key"]], 3, -1)
            zone_wrow = zone_wrow+1
    
    if not priorPeriodReportEnable:
        for zoneCountDataYear in allZoneCountPastYear:
            if zoneCountDataYear["key"] in zone_list.keys():
                zoneWiseCount.cell(zone_list[zoneCountDataYear["key"]], 6, zoneCountDataYear["doc_count"])
                zoneWiseCountYear = zoneCountDataYear["doc_count"]
                change_percent_year = get_percentage_diff(zone_count_list[zoneCountDataYear["key"]],zoneWiseCountYear)
                zoneWiseCount.cell(zone_list[zoneCountDataYear["key"]], 5, change_percent_year/100)
                if zoneCountDataYear["key"] == maxzone_current:
                    maxzone_year = zoneCountDataYear["doc_count"]
            else:
                zone_list.update({zoneCountDataYear["key"]:zone_wrow+1})
                zoneWiseCount.cell(zone_wrow+1, 1, zoneCountDataYear["key"])
                zoneWiseCount.cell(zone_wrow+1, 6, zoneCountDataYear["doc_count"])
                zone_count_list.update({zoneCountDataYear["key"]:zoneCountDataYear["doc_count"]})
                zoneWiseCount.cell(zone_list[zoneCountDataYear["key"]], 5, -1)
                zone_wrow = zone_wrow+1   
        SPT_1="Selected Period Traffic  ( "+str(dateWiseFromDate.strftime('%Y-%m-%d'))+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
        PWT_1="Prior Week Traffic   ( "+str(datePastWeekFromDate.strftime('%Y-%m-%d'))+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
        PYT_1="Prior Year Traffic   ( "+str(datePastYearFromDate.strftime('%Y-%m-%d'))+" TO "+str(datePastYearToDate.strftime('%Y-%m-%d'))+" )"
        zoneWiseCount.cell(1, 2, SPT_1)
        zoneWiseCount.cell(1, 4, PWT_1)
        zoneWiseCount.cell(1, 6, PYT_1)
    else:
        SPT_1="Selected Period Traffic  ( "+str(dateWiseFromDate.strftime('%Y-%m-%d'))+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
        PWT_1="Prior Period Traffic    ( "+str(datePastWeekFromDate.strftime('%Y-%m-%d'))+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
        zoneWiseCount.cell(1, 2, SPT_1)
        zoneWiseCount.cell(1, 4, PWT_1)
    if(maxzone_week == 0):
        maxzoneperct_week = 100
    else:
        maxzoneperct_week = get_percentage_diff(zone_count_list[maxzone_current],maxzone_week)
    footfall_dash.cell(zoned_row+3, 6, str(maxzoneperct_week)+"%")
    if not priorPeriodReportEnable:     
        if(maxzone_year == 0):
            maxzoneperct_year = 100
        else:
            maxzoneperct_year = get_percentage_diff(zone_count_list[maxzone_current],maxzone_year)
        footfall_dash.cell(zoned_row+5, 6, str(maxzoneperct_year)+"%")
    
    footfall_dash.cell(zoned_row+1, 6,zone_count_list[maxzone_current])
    footfall_dash.cell(zoned_row+6, 4, maxzone_current)
    
     
    if(numberOfDayDelta == 0):
        average_totalcount_current = allZoneCountCDate
        average_totalcount_week = allZonePastWCount
        if not priorPeriodReportEnable:
            average_totalcount_year = allZonePastYCount
    else:
        average_totalcount_current = allZoneCountCDate/numberOfDayDelta
        average_totalcount_week = allZonePastWCount/numberOfDayDelta
        if not priorPeriodReportEnable:
            average_totalcount_year = allZonePastYCount/numberOfDayDelta
    footfall_dash.cell(zoned_row+1, 4, int(average_totalcount_current))
     
    diff_percent_avg_countweek = get_percentage_diff(average_totalcount_current,average_totalcount_week)
    footfall_dash.cell(zoned_row+3, 4, str(diff_percent_avg_countweek)+"%")
    if not priorPeriodReportEnable:
        diff_percent_avg_countyear = get_percentage_diff(average_totalcount_current,average_totalcount_year)
        footfall_dash.cell(zoned_row+5, 4, str(diff_percent_avg_countyear)+"%")        
    
    average_hourly_count_current = average_totalcount_current/15
    footfall_dash.cell(zoned_row+1, 5, int(average_hourly_count_current))
    
    average_hourly_count_week = average_totalcount_week/15
    diff_percent_avg_hrcountweek = get_percentage_diff(average_hourly_count_current,average_hourly_count_week)
    footfall_dash.cell(zoned_row+3, 5, str(diff_percent_avg_hrcountweek)+"%")
    if not priorPeriodReportEnable:
        average_hourly_count_year = average_totalcount_year/15
        diff_percent_avg_hrcountyear = get_percentage_diff(average_hourly_count_current,average_hourly_count_year)
        footfall_dash.cell(zoned_row+5, 5, str(diff_percent_avg_hrcountyear)+"%") 
        
if __name__=='__main__':
    if sys.argv[1] is not None:
        data=json.loads(str(sys.argv[1]))
        #print(data)
        site_id = data["site_id"]
        site_name = data["site_name"]
        f_date= datetime. strptime(data["from_date"], '%Y-%m-%d %H:%M:%S')
        t_date = datetime.strptime(data["to_date"], '%Y-%m-%d %H:%M:%S')
        f = f_date
        t=t_date
        zones = data["zones"]
        dayDiffernce = t - f
        numberOfDayDelta=dayDiffernce.days
        fromTimeOfSite=f.time()
        toTimeOfSite=t.time()
        dayTime=toTimeOfSite.hour-fromTimeOfSite.hour
        t2=t
        if (dayTime==0):
            t2 = t - timedelta(days=1)
        if(dayTime>=0):
           pass
        else: 
            numberOfDayDelta=numberOfDayDelta+1
            t2 = t - timedelta(days=1)
        if(numberOfDayDelta>7):
            priorPeriodReportEnable=True
        #print("Zone in input is {}".format(zones))
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/footfall_template.xlsx")
            if priorPeriodReportEnable:
                report_wb = load_workbook("./requests/footfall_template_prior_period.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            #Open Footfall dashboard sheet and update the site title and dates
            #Inorbit Mall (Vadodara) - Footfall Analytics Report 
            #(for the period 1 Jan'21 - 31 Jan'21)
            footfall_dash = report_wb['Summary']
            footfall_dash.cell(3, 3, "{} - Footfall Analytics Report \n (for the period {} - {})".format(site_name,f.strftime("%d %b'%y"),t2.strftime("%d %b'%y")))
            s_no = 1
            row=2
            
            '''while f_date <= t_date:
                print(f_date)
                print(t_date)
                f_date = f_date + timedelta(days=1)'''
            
            #while f_date <= t_date:
            #for i in range(f_date.day, t_date.day+1):            
                # open Footfall Data sheet. 
            footfall_sheet = report_wb['Footfall Data']
            from_date=f_date+timedelta(hours=00)
            to_date=f_date+timedelta(hours=23)
            to_date=to_date+timedelta(minutes=59)
            to_date=to_date+timedelta(seconds=59)
            zonesId = [o["id"] for o in zones]
            #Edited by Sandeep
            mothly_weeklyData(site_id, f_date, t_date)
           
            #Edited by Shubham
            zone_wise_details(site_id, f_date, t_date)
            #Edited by Mukesh
            
            # month_week_payload = month_week_create_request_for_all_data(site_id, zonesId, f_date, t_date)
            # response = requests.get(url+"0",json=month_week_payload,headers=headers,timeout=(300,300))
            # month_week_result=response.json()
            # weekDaysCount(month_week_result)
            # monthDaysCount(month_week_result)
            footfall_payload = create_request_for_all_data(site_id, zonesId, f_date, t_date)
            response = requests.get(url+"0",json=footfall_payload,headers=headers,timeout=(300,300),verify=False)
            footfall_result=response.json()
            #System Health
            system_Health_sheet = report_wb['System_Health']
            system_Health_payload = create_request_for_systemHealth(site_id, f_date, t_date)
            #print("----",system_Health_payload)
            response = requests.get(url+"9000",json=system_Health_payload,headers=headers,timeout=(300,300),verify=False)
            system_Health_result=response.json()
            #print(system_Health_result)
            dateInterval=footfall_result["aggregations"]["date_interval"]["buckets"]
            for dateWisetData in dateInterval:
                allZoneCountDetails=dateWisetData["ZoneDetails"]["buckets"]
                for zoneCountData in allZoneCountDetails:
                    footfall_sheet.cell(row, 1, s_no)
                    footfall_sheet.cell(row, 2, dateWisetData["key_as_string"].split("T")[0])
                    footfall_sheet.cell(row, 3, zoneCountData['key'])
                    footfall_sheet.cell(row, 4, zoneCountData["doc_count"])
                    genderData=zoneCountData['ZoneGender']['buckets']
                    for gender in genderData:
                        if gender["key"] == "Male":
                            footfall_sheet.cell(row, 5, gender["doc_count"])
                        if gender["key"] == "Female":
                            footfall_sheet.cell(row, 6, gender["doc_count"])
                    
                    ageGroupData=zoneCountData['ZoneAgeGroup']['buckets']
                    for ageData in ageGroupData:
                        if ageData["key"] == "0-15":
                            footfall_sheet.cell(row, 7, ageData["doc_count"])
                        if ageData["key"] == "16-35":
                            footfall_sheet.cell(row, 8, ageData["doc_count"])
                        if ageData["key"] == "16-25":
                            footfall_sheet.cell(row, 9, ageData["doc_count"])
                        if ageData["key"] == "26-35":
                            footfall_sheet.cell(row, 10, ageData["doc_count"])
                        if ageData["key"] == "36-55":
                            footfall_sheet.cell(row, 11, ageData["doc_count"])
                        if ageData["key"] == "55+":
                            footfall_sheet.cell(row, 12, ageData["doc_count"])
                    row=row+1
                    s_no=s_no+1
                
                f_date = f_date + timedelta(days=1)
            
            footfall_Hourly_sheet = report_wb['Footfall_Hourly_Data']
            allZoneCountCurrentDate=footfall_result["aggregations"]["Zone_wise_total"]["buckets"]
            # print("Len wise total count -- ",len(allZoneCountCurrentDate))
            locationWiseAgeGenderDic={}
            locationCol=4
            for zoneCountData in allZoneCountCurrentDate:
                footfall_Hourly_sheet.cell(2, locationCol, zoneCountData["key"])
                footfall_Hourly_sheet.cell(3, locationCol, 'Total')
                footfall_Hourly_sheet.cell(3, locationCol+1, 'Male')
                footfall_Hourly_sheet.cell(3, locationCol+2, 'Female')
                footfall_Hourly_sheet.cell(3, locationCol+3, 'Age Group (0-15)')
                footfall_Hourly_sheet.cell(3, locationCol+4, 'Age Group (16-25)')
                footfall_Hourly_sheet.cell(3, locationCol+5, 'Age Group (26-35)')
                footfall_Hourly_sheet.cell(3, locationCol+6, 'Age Group (36-55)')
                footfall_Hourly_sheet.cell(3, locationCol+7, 'Age Group (55+)')
                locationWiseAgeGenderDic[zoneCountData["key"]]=locationCol
                locationCol=locationCol+8
            hr_col=4
            new_hours_data=f
            hr_row=4
            time_row_dic={}
            while new_hours_data < t:
                time_row_dic[new_hours_data]=hr_row
                footfall_Hourly_sheet.cell(hr_row, 2, new_hours_data)
                new_hours_data = new_hours_data+ timedelta(hours=1)
                footfall_Hourly_sheet.cell(hr_row, 3, new_hours_data)
                hr_row=hr_row + 1
            
            for i in range(4,hr_row):
                for j in range(hr_col,hr_col+len(allZoneCountCurrentDate*8)):
                    footfall_Hourly_sheet.cell(i, j, 0)
            # print("Dic",locationWiseAgeGenderDic)
            # print("time_row_dic",time_row_dic)
            HoulyFootfallCountDetails=footfall_result["aggregations"]["HoulyFootfallCountDetails"]["buckets"]
            for each_hour_data in HoulyFootfallCountDetails:
                tm=datetime.strptime(each_hour_data["key_as_string"], '%Y-%m-%d %H:%M:%S')
                if (tm in time_row_dic):
                    # ZoneDetails_Hourly_Count=footfall_result["aggregations"]["ZoneDetails_Hourly_Count"]["buckets"]
                    ZoneDetails_Hourly_Count=each_hour_data["ZoneDetails_Hourly_Count"]["buckets"]
                    for ZonesData in ZoneDetails_Hourly_Count:
                        zoneColNo=locationWiseAgeGenderDic[ZonesData['key']]
                        # footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, ZonesData['key'])
                        footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo , ZonesData["doc_count"])
                        # print("--------ZonesData['key']", ZonesData['key'],ZonesData["doc_count"])
                        GenderHourly=ZonesData['GenderHourly']['buckets']
                        for gender in GenderHourly:
                            if gender["key"] == "Male":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+1, gender["doc_count"])
                            if gender["key"] == "Female":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+2, gender["doc_count"])
                        AgeGroupHourly=ZonesData['AgeGroupHourly']['buckets']
                        for ageData in AgeGroupHourly:
                            if ageData["key"] == "0-15":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+3, ageData["doc_count"])
                            if ageData["key"] == "16-25":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+4, ageData["doc_count"])
                            if ageData["key"] == "26-35":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+5, ageData["doc_count"])
                            if ageData["key"] == "36-55":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+6, ageData["doc_count"])
                            if ageData["key"] == "55+":
                                footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+7, ageData["doc_count"])
                    
                    # footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, each_hour_data["doc_count"])
                    # totalCount=totalCount+each_hour_data["doc_count"]
                    # genderData=zoneCountData['ZoneGender']['buckets']
                    # for gender in genderData:
                    #     if gender["key"] == "Male":
                    #         footfall_sheet.cell(row, 5, gender["doc_count"])
                    #     if gender["key"] == "Female":
                    #         footfall_sheet.cell(row, 6, gender["doc_count"])
                    
                    # ageGroupData=zoneCountData['ZoneAgeGroup']['buckets']
                    # for ageData in ageGroupData:
                    #     if ageData["key"] == "0-15":
                    #         footfall_sheet.cell(row, 7, ageData["doc_count"])
                    #     if ageData["key"] == "16-35":
                    #         footfall_sheet.cell(row, 8, ageData["doc_count"])
                    #     if ageData["key"] == "16-25":
                    #         footfall_sheet.cell(row, 9, ageData["doc_count"])
                    #     if ageData["key"] == "26-35":
                    #         footfall_sheet.cell(row, 10, ageData["doc_count"])
                    #     if ageData["key"] == "36-55":
                    #         footfall_sheet.cell(row, 11, ageData["doc_count"])
                    #     if ageData["key"] == "55+":
                    #         footfall_sheet.cell(row, 12, ageData["doc_count"])
            #For Getting Hourly Footfall data
            '''footfall_Hourly_sheet = report_wb['Footfall_Hourly_Data']
            ZoneDetails_Hourly_Count=footfall_result["aggregations"]["ZoneDetails_Hourly_Count"]["buckets"]
            hr_col=4
            new_hours_data=f
            hr_row=4
            time_row_dic={}
            while new_hours_data < t:
                time_row_dic[new_hours_data]=hr_row
                footfall_Hourly_sheet.cell(hr_row, 2, new_hours_data)
                new_hours_data = new_hours_data+ timedelta(hours=1)
                footfall_Hourly_sheet.cell(hr_row, 3, new_hours_data)
                hr_row=hr_row + 1
            footfall_Hourly_sheet.cell(hr_row, 3, "Total Count")
            for i in range(4,hr_row):
                for j in range(hr_col,hr_col+len(ZoneDetails_Hourly_Count)):
                    footfall_Hourly_sheet.cell(i, j, 0)
            for ZonesData in ZoneDetails_Hourly_Count:
                footfall_Hourly_sheet.cell(3, hr_col, ZonesData['key'])
                hourly_count_data=ZonesData['HoulyCount']['buckets']
                totalCount=0
                for each_hour_data in hourly_count_data:
                    tm=datetime.strptime(each_hour_data["key_as_string"], '%Y-%m-%d %H:%M:%S')
                    if (tm in time_row_dic):
                        footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, each_hour_data["doc_count"])
                        totalCount=totalCount+each_hour_data["doc_count"]
                footfall_Hourly_sheet.cell(hr_row, hr_col, totalCount)
                hr_col=hr_col+1'''
            
            '''#Add new logic for write dress color data for age wise and gender wise
            age_gender_row={"Male":2,"Female":3,"0-15":4,"16-25":5,"26-35":6,"36-55":7,"55+":8}
            # color_coloum={"Red":2,"Blue":3,"Green":4,"Yellow":5,"Orange":6,"Purple":7,"Pink":8,"Brown":9,"White":10,"Gray":11,"Cyan":12,"Black":13,"SlateGray":14}
            color_coloum={}
            dress_color_sheet = report_wb['DressColor']
            gender_dress_color_data=footfall_result["aggregations"]["GenderWiseDressColor"]["buckets"]
            age_dress_color_data=footfall_result["aggregations"]["AgeWiseDressColor"]["buckets"]
            # print(gender_dress_color_data)
            # print(age_dress_color_data)
            cgNumber=2
            for g_dress_data in gender_dress_color_data:
                gender_row_number=g_dress_data['key']
                dress_color_data=g_dress_data['DressColor']['buckets']
                for color_counts_data in dress_color_data:
                    g_colorName=color_counts_data['key']
                    if g_colorName not in color_coloum:
                        color_coloum[g_colorName]=cgNumber
                        cgNumber=cgNumber+1
                        # g_colorName="Other"
                    
                    g_colorCountValue=color_counts_data['doc_count']
                    dress_color_sheet.cell(age_gender_row[gender_row_number], color_coloum[g_colorName], g_colorCountValue)
            
            for age_dress_data in age_dress_color_data:
                age_row_number=age_dress_data['key']
                dress_color_data=age_dress_data['DressColor']['buckets']
                for color_counts_data in dress_color_data:
                    a_colorName=color_counts_data['key']
                    if a_colorName not in color_coloum:
                        color_coloum[a_colorName]=cgNumber
                        cgNumber=cgNumber+1
                        # a_colorName="Other"
                    a_colorCountValue=color_counts_data['doc_count']
                    dress_color_sheet.cell(age_gender_row[age_row_number], color_coloum[a_colorName], a_colorCountValue)'''
            
            dress_color_sheet = report_wb['DressColor']
            #Dress ORG color
            dress_color_data=footfall_result["aggregations"]["DressColor"]["buckets"]
            dc_r_number=11
            for color_counts_data in dress_color_data:
                colorName=color_counts_data['key']
                colorCountValue=color_counts_data['doc_count']
                dress_color_sheet.cell( dc_r_number,1,(dc_r_number-10))
                dress_color_sheet.cell( dc_r_number,2,colorName)
                dress_color_sheet.cell( dc_r_number,3, colorCountValue)
                dc_r_number=dc_r_number+1
            
            dress_color_sheet.cell( dc_r_number,1,(dc_r_number-10))
            dress_color_sheet.cell( dc_r_number,2,"Other Colors")
            sum_other_doc_count=footfall_result["aggregations"]["DressColor"]["sum_other_doc_count"]
            dress_color_sheet.cell( dc_r_number,3, sum_other_doc_count)
            #Dress color Group
            dress_color_data=footfall_result["aggregations"]["DressColorGroup"]["buckets"]
            dc_c_number=3
            for color_counts_data in dress_color_data:
                g_colorName=color_counts_data['key']
                g_colorCountValue=color_counts_data['doc_count']
                dress_color_sheet.cell( 4,dc_c_number,g_colorName)
                dress_color_sheet.cell( 5,dc_c_number, g_colorCountValue)
                dc_c_number=dc_c_number+1
            
        
        
        #print("system_Health_result-----",system_Health_result)
        sh_data=system_Health_result["hits"]["hits"]
        sh_no=1
        for sh_status in sh_data:
            system_Health_sheet.cell(sh_no+1, 1, sh_no)
            system_Health_sheet.cell(sh_no+1, 2, sh_status['_source']['DateTime'])
            system_Health_sheet.cell(sh_no+1, 3, sh_status['_source']['camera_name'])
            system_Health_sheet.cell(sh_no+1, 4, sh_status['_source']['camera_status'])
            system_Health_sheet.cell(sh_no+1, 5, sh_status['_source']['is_resolution_change'])
            system_Health_sheet.cell(sh_no+1, 6, sh_status['_source']['resolution'])
            sh_no=sh_no+1            
        site_name=site_name.replace(" ", "-")  
        ed_date= datetime. strptime(data["fileName"], '%Y-%m-%d')   
        #ed_date.strftime("%d-%m-%Y")    
        report_wb.save('./report/Smarg_{}_Footfall_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
        print('Smarg_{}_Footfall_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
    
        
