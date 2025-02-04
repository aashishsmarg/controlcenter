import json
import os
import sys
import requests
# Writing to an excel  
# sheet using Python 
import xlwt
import xlrd
from xlwt import Workbook
from openpyxl import load_workbook
from datetime import datetime,timedelta
#from pandas import json_normalize 
PC_REQUEST_PATH = "./requests/event_request.json"
date=None
url = "http://smargtech.in:36523/filebeat-*/_search/?size="
headers = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzpzbWFyZzEyM3NtYXJnIw=='}
def create_request_for_all_data(site_id, zone_ids, from_date, to_date,event_name):
    json_obj=None
    with open(PC_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][0]["term"]["SiteId"]=site_id
    json_obj["query"]["bool"]["filter"][1]["terms"]["ZoneId"]=zone_ids
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    json_obj["query"]["bool"]["filter"][6]["term"]["Event.keyword"]=event_name
    return json_obj
def getEvent_VS_TotalCountData(site_id, zone_ids, from_date, to_date):
    Event_VS_Total_sheet = report_wb['Event_VS_Total']
    fromDate="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    todate="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    
    
    jsonRequestData={"query":{"bool":{"must":[],"filter":[
        {"term":{"SiteId":site_id}},{"terms":{"ZoneId":zone_ids}},{"term":{"isViolation":0}},{"term":{"Direction.keyword":"IN"}},{"term":{"NotificationType":"7"}},
        {"range":{"DateTime":{"gte":fromDate,"lte":todate}}}]}},
        "aggs":{"TotalCountEvent":{"date_histogram":{"field":"DateTime","interval": "day"},"aggs":{"EventData":{"terms":{"field":"Event.keyword"}}}}}}
    eventresponse = requests.get(url+"0",json=jsonRequestData,headers=headers,timeout=(300,300))
    footfall_result=eventresponse.json()
    Event_VS_Total_Count=footfall_result["aggregations"]["TotalCountEvent"]['buckets']
    evRow=2
    evCol=2
    eventDic={}
    Event_VS_Total_sheet.cell(1, 1, "Date")
    for EventBucket in Event_VS_Total_Count:
        e_tm=datetime. strptime(EventBucket["key_as_string"], '%Y-%m-%dT%H:%M:%S.%fZ')
        Event_VS_Total_sheet.cell(evRow, 1, e_tm)
        event_data=EventBucket["EventData"]['buckets']
        for perEventData in event_data:
            eventName=perEventData['key']
            if eventName=="No Event":
                eventName="Total Footfall"
            personCount=perEventData['doc_count'] 
            if (eventName in eventDic):
                Event_VS_Total_sheet.cell(evRow, eventDic[eventName], personCount)
            else:
                Event_VS_Total_sheet.cell(1, evCol, eventName)
                Event_VS_Total_sheet.cell(evRow, evCol, personCount)
                eventDic[eventName]=evCol
                evCol=evCol+1
        evRow=evRow+1
if __name__=='__main__':
    if sys.argv[1] is not None:
        data=json.loads(str(sys.argv[1]))
        site_id = data["site_id"]
        site_name = data["site_name"]
        event_name =data["event_name"]
        f_date= datetime. strptime(data["from_date"], '%Y-%m-%d %H:%M:%S')
        t_date = datetime.strptime(data["to_date"], '%Y-%m-%d %H:%M:%S')
        f = f_date
        t=t_date
        zones = data["zones"]
        #print("Zone in input is {}".format(zones))
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/event_report_template.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            #Open Footfall dashboard sheet and update the site title and dates
            #Inorbit Mall (Vadodara) - Footfall Analytics Report 
            #(for the period 1 Jan'21 - 31 Jan'21)
            footfall_dash = report_wb['Dashboard']
            footfall_dash.cell(2, 2, "{} - {}- Event Analytics Report \n (for the period {} - {})".format(site_name,event_name,f.strftime("%d %b'%y"),t.strftime("%d %b'%y")))
            s_no = 1
            row=2
            
            '''while f_date <= t_date:
                print(f_date)
                print(t_date)
                f_date = f_date + timedelta(days=1)'''
            #while f_date <= t_date:
            #for i in range(f_date.day, t_date.day+1):            
                # open Footfall Data sheet. 
            footfall_sheet = report_wb['Event Data']
            from_date=f_date+timedelta(hours=00)
            to_date=f_date+timedelta(hours=23)
            to_date=to_date+timedelta(minutes=59)
            to_date=to_date+timedelta(seconds=59)
            zonesId = [o["id"] for o in zones]
            getEvent_VS_TotalCountData(site_id, zonesId, f_date, t_date)
            footfall_payload = create_request_for_all_data(site_id, zonesId, f_date, t_date,event_name)
            response = requests.get(url+"0",json=footfall_payload,headers=headers,timeout=(300,300))
            footfall_result=response.json()
            #print(footfall_result)
            dateInterval=footfall_result["aggregations"]["date_interval"]["buckets"]
            for dateWisetData in dateInterval:
                allZoneCountDetails=dateWisetData["ZoneDetails"]["buckets"]
                for zoneCountData in allZoneCountDetails:
                    footfall_sheet.cell(row, 1, s_no)
                    footfall_sheet.cell(row, 2, dateWisetData["key_as_string"].split("T")[0])
                    footfall_sheet.cell(row, 3, zoneCountData['key'])
                    footfall_sheet.cell(row, 4, zoneCountData["doc_count"])
                    genderData=zoneCountData['ZoneGender']['buckets']
                    for gender in genderData:
                        if gender["key"] == "Male":
                            footfall_sheet.cell(row, 5, gender["doc_count"])
                        if gender["key"] == "Female":
                            footfall_sheet.cell(row, 6, gender["doc_count"])
                    
                    # ageGroupData=zoneCountData['ZoneAgeGroup']['buckets']
                    # for ageData in ageGroupData:
                    #     if ageData["key"] == "0-15":
                    #         footfall_sheet.cell(row, 7, ageData["doc_count"])
                    #     if ageData["key"] == "16-35":
                    #         footfall_sheet.cell(row, 8, ageData["doc_count"])
                    #     if ageData["key"] == "36-55":
                    #         footfall_sheet.cell(row, 9, ageData["doc_count"])
                    #     if ageData["key"] == "55+":
                    #         footfall_sheet.cell(row, 10, ageData["doc_count"])
                    row=row+1
                    s_no=s_no+1
                
                f_date = f_date + timedelta(days=1)
            #For Getting Hourly Footfall data
            footfall_Hourly_sheet = report_wb['Event_Hourly_Data']
            ZoneDetails_Hourly_Count=footfall_result["aggregations"]["ZoneDetails_Hourly_Count"]["buckets"]
            hr_col=4
            new_hours_data=f
            hr_row=4
            time_row_dic={}
            while new_hours_data < t:
                time_row_dic[new_hours_data]=hr_row
                footfall_Hourly_sheet.cell(hr_row, 2, new_hours_data)
                new_hours_data = new_hours_data+ timedelta(hours=1)
                footfall_Hourly_sheet.cell(hr_row, 3, new_hours_data)
                hr_row=hr_row + 1
            footfall_Hourly_sheet.cell(hr_row, 3, "Total Count")
            for i in range(4,hr_row):
                for j in range(hr_col,hr_col+len(ZoneDetails_Hourly_Count)):
                    footfall_Hourly_sheet.cell(i, j, 0)
            for ZonesData in ZoneDetails_Hourly_Count:
                footfall_Hourly_sheet.cell(3, hr_col, ZonesData['key'])
                hourly_count_data=ZonesData['HoulyCount']['buckets']
                totalCount=0
                for each_hour_data in hourly_count_data:
                    tm=datetime. strptime(each_hour_data["key_as_string"], '%Y-%m-%d %H:%M:%S')
                    if (tm in time_row_dic):
                        footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, each_hour_data["doc_count"])
                        totalCount=totalCount+each_hour_data["doc_count"]
                footfall_Hourly_sheet.cell(hr_row, hr_col, totalCount)
                hr_col=hr_col+1
        site_name=site_name.replace(" ", "-")  
        ed_date= datetime. strptime(data["fileName"], '%Y-%m-%d')   
        #ed_date.strftime("%d-%m-%Y")    
        report_wb.save('./report/Smarg_{}_Event_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
        print("Report Generate successfull")
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
