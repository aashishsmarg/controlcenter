import json
import os
import sys
import requests
# Writing to an excel  
# sheet using Python 
from openpyxl import load_workbook
from datetime import datetime,timedelta
VIOLATION_REQUEST_PATH = "./requests/violation_request.json"
date=None

# vc_url="http://localhost:9200/filebeat-*/_search/?scroll=1m"
url = "http://smargtech.in:36523/filebeat-*/_search/?scroll=1m"
#url = "http://178.128.249.67:36523/filebeat-*/_search/?size=1000"
#url = "http://44.235.236.183:9200/filebeat-*/_search/?size=1000"
headers = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzpzbWFyZzEyMw=='}
# headers_vc = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzpzbWFyZzEyM3NtYXJnIw=='}
def create_violation_request(site_id, zone_id, from_date, to_date, notification_type):
    #json=json.loads("{\"query\":{\"bool\":{\"must\":[],\"filter\": [{\"term\": {\"SiteId\": 111}},{\"term\": {\"ZoneId\": 1}},{\"term"\: {\"isViolation\": 0}},{\"term\": {\"Direction.keyword\": \"IN\"}},{\"term\": {\"NotificationType\": \"7\"}},{\"range\": {\"DateTime\": {\"gte\": \"2020-09-26T00:00:00Z\",\"lte\": \"2020-12-26T11:00:00Z\",}}},],}},\"sort\": [{\"@timestamp\": {\"order\": \"desc\"}}],}")
    json_obj=None
    with open(VIOLATION_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][0]["term"]["SiteId"]=site_id
    #json_obj["query"]["bool"]["filter"][1]["terms"]["ZoneId"] = zone_id
    json_obj["query"]["bool"]["filter"][2]["terms"]["NotificationType"] = notification_type    
    json_obj["query"]["bool"]["filter"][3]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][3]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    print("request is {}".format(json_obj))
    return json_obj
if __name__=='__main__':
    if sys.argv[1] is not None:
        print("Json string in request is: {}".format(sys.argv[1]))
        data=json.loads(str(sys.argv[1]))
        site_id = data["site_id"]
        site_name = data["site_name"]
        f_date= datetime. strptime(data["from_date"], '%Y-%m-%d %H:%M:%S')
        t_date = datetime.strptime(data["to_date"], '%Y-%m-%d %H:%M:%S')
        f = f_date
        t=t_date
        zones = data["zones"]
        #print("Zone in input is {}".format(zones))
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/violation_template.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            s_no = 1
            row=2
            violation_col_dic={}
            violation_col=4
            for i in range(f_date.day, t_date.day+1):            
                # open Violation Data sheet.
                violation_sheet = report_wb['Violation Data']
                from_date=f_date+timedelta(hours=00)
                to_date=f_date+timedelta(hours=23)
                to_date=to_date+timedelta(minutes=59)
                to_date=to_date+timedelta(seconds=59)
                zonesId = [o["id"] for o in zones]
                notification_type=["1","2","22","23"]
                violation_payload = create_violation_request(site_id, zonesId, from_date, to_date,notification_type)
                response = requests.get(url,json=violation_payload,headers=headers,timeout=(300,300))
                violation_result=response.json()
                print(violation_result)
                allZoneCountDetails=violation_result["aggregations"]["ZoneDetails"]["buckets"]
                allViolationName=violation_result["aggregations"]["TotalViolation"]["buckets"]
                
                for ViolationName in allViolationName:
                    if ViolationName["key"] not in violation_col_dic:
                        violation_col_dic[ViolationName["key"]]=violation_col
                        violation_sheet.cell(1, violation_col, ViolationName["key"])
                        violation_col=violation_col+1
                        
                for zoneCountData in allZoneCountDetails:
                    #print(zoneCountData['key'])
                    #print(zoneCountData['doc_count'])
                    #Fetch Violation data
                    #Fill up violation sheet for specific date and zone
                    violation_sheet.cell(row, 1, s_no)
                    violation_sheet.cell(row, 2, "{}".format(f_date.strftime("%m/%d/%Y")))
                    violation_sheet.cell(row, 3, zoneCountData['key'])
                    notificationData=zoneCountData['ZoneNotification']['buckets']
                    for notificationType in notificationData:
                        violation_sheet.cell(row, violation_col_dic[notificationType["key"]], notificationType["doc_count"])
                        # if(notificationType["key"]== 22):
                        #     violation_sheet.cell(row, 4, notificationType["doc_count"])
                        # if(notificationType["key"]== 23):
                        #     violation_sheet.cell(row, 5, notificationType["doc_count"])
                        # if(notificationType["key"]== 1):
                        #     violation_sheet.cell(row, 6, notificationType["doc_count"])
                        # if(notificationType["key"]== 2):
                        #     violation_sheet.cell(row, 7, notificationType["doc_count"])
                    row=row+1
                    s_no=s_no+1
                
                f_date = f_date + timedelta(days=1)
        site_name=site_name.replace(" ", "-")       
        report_wb.save('./report/Smarg_{}_Violation_Analytics_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),t.strftime("%d-%m-%Y")))
        print("Report Generate successfull")
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
    
