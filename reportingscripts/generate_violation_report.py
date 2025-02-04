import json
import os
import sys
import requests
# Writing to an excel  
# sheet using Python 
from PIL import Image
from openpyxl import load_workbook,drawing
from datetime import datetime,timedelta
VIOLATION_REQUEST_PATH = "./requests/violation_request.json"
ViolationImagePath='https://smargtechnologies.in:3015/surveillance/'
date=None
url = "http://localhost:3021/surveillance/notification?"
#url = "http://44.235.236.183:9200/filebeat-*/_search/?scroll=1m"
if __name__=='__main__':
    if sys.argv[1] is not None:
        data=json.loads(str(sys.argv[1]))
        site_name = data["site_name"]
        site_id = data["site_id"]
        f_date= datetime. strptime(data["from_date"], '%Y-%m-%d %H:%M:%S')
        t_date = datetime.strptime(data["to_date"], '%Y-%m-%d %H:%M:%S')
        f = f_date
        t=t_date
        
        delta = t_date - f_date
        print(delta.days)
        auth_token= data["auth_token"]
        #print("Zone in input is {}".format(zones))
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/Violations_Report_Image.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            s_no = 1
            row=10         
            # open Violation Data sheet.
            violation_sheet = report_wb['Violation Data']
            query = {'from':f_date, 'to':t_date,'site_id':site_id ,'is_report':1}
            headers = {'content_type': 'application/json','auth_token': auth_token}
            response = requests.get(url,headers=headers,params=query,timeout=(300,300))
            
            now = datetime.now()
            dt_string = now.strftime("%Y-%m-%d")
            violation_sheet.cell(3, 5, site_name)
            violation_sheet.cell(4, 5, dt_string)
            violation_sheet.cell(5, 5, f_date)
            violation_sheet.cell(6, 5, t_date)
            
            violations_result=response.json()
            #print("violations_result   : ",violations_result)
            if 'result' in violations_result:
                violations_result=violations_result['result']
                violation_sheet.cell(7, 5, len(violations_result))
                for violationData in violations_result:
                    violation_sheet.cell(row, 2, s_no)
                    imageName=''
                    try:
                        split_string = violationData['image_url_1'].split('/')
                        imageName=split_string[-1]
                    except Exception as e1:
                        imageName=violationData['image_url_1']
                    path=ViolationImagePath+''+violationData['image_url_1']
                    c1=violation_sheet.cell(row, 3, imageName)
                    c1.hyperlink=path
                    violation_sheet.cell(row, 4, violationData['notification_type'])
                    violation_sheet.cell(row, 5, violationData['date'])
                    violation_sheet.cell(row, 6, violationData['camera_name'])
                    violation_sheet.cell(row, 7, violationData['status'])
                    violation_sheet.cell(row, 8, violationData['discription'])
                    violation_sheet.cell(row, 9, violationData['updated_by'])
                    violation_sheet.cell(row, 10, violationData['updated_at'])
                    violation_sheet.cell(row, 11, violationData['comment'])
                    row=row+1
                    s_no=s_no+1
        site_name=site_name.replace(" ", "-") 
        ed_date= datetime. strptime(data["fileName"], '%Y-%m-%d')
        report_wb.save('./report/Smarg_{}_Violation_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
        print("Report Generate successfull")
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
    
