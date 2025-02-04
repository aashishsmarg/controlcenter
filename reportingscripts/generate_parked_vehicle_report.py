import json
import sys
import requests
import time
# Writing to an excel  
# sheet using Python 
from PIL import Image
from openpyxl import load_workbook,drawing
from datetime import datetime,timedelta
# from bunch import bunchify
VC_REQUEST_PATH = "./requests/total_vehicle_count_4W.json"
date=None 
vc_url="http://localhost:9200/filebeat-*/_search/?size=0"
vc_elk_url="http://localhost:9200/filebeat-*/_search/?filter_path=hits.hits._source"

#vc_url="http://smargtech.in:36523/filebeat-*/_search/?size=0"
#vc_elk_url="http://smargtech.in:36523/filebeat-*/_search/?filter_path=hits.hits._source"
    
headers_vc = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzpzbWFyZzEyM3NtYXJnIw=='}
# url = "http://smargtech.in:36523/surveillance/vehicle_list_anpr?"
VehicleImagePath='/opt/lampp/htdocs/Surveillanceimage/'
isDebug=True
isimageRequired = False
def setMaxSize() :
    url = "http://localhost:9200/filebeat-*/_settings"
    payload = json.dumps({"index": {"max_result_window": 210000}})
    headers = {'Authorization': 'Basic ZWxhc3RpYzpzbWFyZzEyM3NtYXJnIw==','Content-Type': 'application/json'}
    response = requests.request("PUT", url, headers=headers, data=payload)

def elkVehicleAPI(site_id, from_date, to_date):
    fr_date="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    t_date="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    finaleRequest = {
            "size": 200000,
            "track_total_hits": True,
            "aggs": { "type_count": { "cardinality": { "field": "ANPR_Marge.keyword" } } },
            "query": { "bool": { "must": { "exists": { "field": "MainEntryCount" } }, "filter": [{ "bool": { "should": [{ "bool": { "must": [ { "match": { "IsCheckOut": False } },{ "match": { "Direction": "IN" } }] } }, { "bool": { "must": [ { "match": { "IsCheckOut": True } },{ "match": { "Direction": "OUT" } }] } }] } }, { "term": { "SiteId": site_id } }, { "range": { "DateTime": { "gte": fr_date, "lte": t_date } } }] } },
            "sort": [{ "DateTime": { "order": "asc", "unmapped_type": "date" } }],
            "_source": ["DateTime","Direction","CheckInTime","ANPR","ANPR_Marge","VehicleNumberPlateImagePath","CheckOutTime","NotificationType","Notification","VehicleType","FirstVisit","VisitCount","Area","LaneIn","CheckOutTime","PaymentMode","ParkingCharge","DwellTime","VehicleCheckInTime"]
        }
    
    py_response = requests.get(vc_elk_url,json=finaleRequest,headers=headers_vc)
    py_vehicle_count_response=py_response.json()
    #newl=len(py_vehicle_count_response["hits"]["hits"])
    
    if (len(py_vehicle_count_response)>0 and len(py_vehicle_count_response["hits"]["hits"]) > 0):
        VehiclesFinaldata = [element['_source'] for element in py_vehicle_count_response["hits"]["hits"]]
        return VehiclesFinaldata
    return []
def verifyValue(data) :
    if (data == None or data == ""  or data == "null" or data == "undefined"):
        return ""
    else :
        formatted_chkin_Date=datetime.strptime(data, '%Y-%m-%dT%H:%M:%SZ')
        created_at = formatted_chkin_Date.strftime("%d-%m-%Y %H:%M:%S")
        return formatted_chkin_Date
def create_request_for_vehicle_count(site_id, from_date, to_date):
    #json=json.loads("{\"query\":{\"bool\":{\"must\":[],\"filter\": [{\"term\": {\"SiteId\": 111}},{\"term\": {\"ZoneId\": 1}},{\"term"\: {\"isViolation\": 0}},{\"term\": {\"Direction.keyword\": \"IN\"}},{\"term\": {\"NotificationType\": \"7\"}},{\"range\": {\"DateTime\": {\"gte\": \"2020-09-26T00:00:00Z\",\"lte\": \"2020-12-26T11:00:00Z\",}}},],}},\"sort\": [{\"@timestamp\": {\"order\": \"desc\"}}],}")
    json_obj=None
    with open(VC_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][0]["term"]["SiteId"]=site_id
    json_obj["query"]["bool"]["filter"][4]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][4]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    return json_obj
def countDateWise(site_id,from_date,to_date):
    try:
        vehicle_count_payload = create_request_for_vehicle_count(site_id,from_date,to_date)
        #print(vehicle_count_payload)
        vehicle_count_response_date_wise = requests.get(vc_url,json=vehicle_count_payload,headers=headers_vc,timeout=(300,300))
        vehicle_count_response_date_wise=vehicle_count_response_date_wise.json()
        # print('vehicle_count_response')
        # print(vehicle_count_response)
        TotalVehicleCount=vehicle_count_response_date_wise['aggregations']['TotalVehicleCount']['buckets'][0]["doc_count"]
        VehicleTypeDetails = vehicle_count_response_date_wise['aggregations']['VehicleType']['buckets']
        ZoneViseEntryDetails = vehicle_count_response_date_wise['aggregations']['ZoneDetails']['buckets']
        vehicleTypeDict = {}
        zoneWiseDict = {}
        
        for vehicleType in VehicleTypeDetails:
            vehicleTypeDict[vehicleType['key']] = vehicleType['doc_count']
        
        for zone in ZoneViseEntryDetails:
            zoneWiseDict[zone['key']]=zone['doc_count']
        f_date = datetime.strftime(from_date, "%d-%m-%Y")      
        return {'TotalVehicleCount':TotalVehicleCount,'VehicleTypeDetails':vehicleTypeDict,'ZoneViseEntryDetails':zoneWiseDict}
    except:
        return {'TotalVehicleCount':0,'VehicleTypeDetails':{},'ZoneViseEntryDetails':{}}
if __name__=='__main__':
    if sys.argv[1] is not None:
        data=json.loads(str(sys.argv[1]))
        setMaxSize()
        onlinePayment=0
        cashPayment=0
        all_date_list =[] 
        totalVehicle_list =[]
        vehicleType_list=[]
        zone_list =[]
        allDataCountDict = {}
        
        site_name = data["site_name"]
        site_id = data["site_id"]
        f_date =data["from_date"]
        t_date =data["to_date"]
        f= datetime. strptime(data["from_date"], '%Y-%m-%dT%H:%M:%S')
        t = datetime.strptime(data["to_date"], '%Y-%m-%dT%H:%M:%S')
        vehicle_count_payload = create_request_for_vehicle_count(site_id,f,t)
        response = requests.get(vc_url,json=vehicle_count_payload,headers=headers_vc,timeout=None)
        vehicle_count_response=response.json()
        
        fromTimeOfSite=f.time()
        toTimeOfSite=t.time()
        fromDateOfSite=f.date()
        toDateOfSite=t.date()
        dayTime=toTimeOfSite.hour-fromTimeOfSite.hour
        isSameDay=False
        if(dayTime>0):
            isSameDay=True
        delta = timedelta(days=1)
        while fromDateOfSite <= toDateOfSite:
            if not isSameDay and fromDateOfSite==toDateOfSite:
                fromDateOfSite +=delta
            else:
                dateWiseFromDate=str(fromDateOfSite)+"T"+str(fromTimeOfSite)
                dateWiseToDate=""
                if isSameDay:
                    dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
                    fromDateOfSite +=delta
                else:
                    fromDateOfSite +=delta
                    dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
                dateWiseFromDate =datetime.strptime(dateWiseFromDate, "%Y-%m-%dT%H:%M:%S")
                dateWiseToDate =datetime.strptime(dateWiseToDate, "%Y-%m-%dT%H:%M:%S")
                print(site_id,dateWiseFromDate,dateWiseToDate)
                allCountsDict = countDateWise(site_id,dateWiseFromDate,dateWiseToDate)
                new_f_date = datetime.strftime(dateWiseFromDate, "%d-%m-%Y")
                allDataCountDict[new_f_date] = allCountsDict 
        auth_token= data["auth_token"]
        #print("Zone in input is {}".format(zones))
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/Parked_Vehicle_Report_Image.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            s_no = 1
                     
            # open Violation Data sheet.
            vehicle_sheet = report_wb['Parked Vehicles Report']
            date_wise_count_sheet = report_wb['Date Wise Count']
            key_list =[]
            zone_header_dict = {}
            for dates in allDataCountDict:
                if 'ZoneViseEntryDetails' in  allDataCountDict[dates]:
                    if len(allDataCountDict[dates]['ZoneViseEntryDetails'])>0:
                        key_list.extend(list(allDataCountDict[dates]['ZoneViseEntryDetails'].keys()))
            keys_list = set(key_list)
            for index, value in enumerate(keys_list):
                zone_header_dict[index+5] = value
            for col_index,value in zone_header_dict.items():
                c8=date_wise_count_sheet.cell(2, col_index, value)
            i=3
            j=5
            for date,allData in allDataCountDict.items():
                totalVehicalCount = allData['TotalVehicleCount']
                a1=date_wise_count_sheet.cell(i, 1, date)
                a2 = date_wise_count_sheet.cell(i, 2, totalVehicalCount)
                if 'Eco' in allData['VehicleTypeDetails']:
                    date_wise_count_sheet.cell(i, 3, allData['VehicleTypeDetails']['Eco'])
                else:
                    date_wise_count_sheet.cell(i, 3, 0)
                if 'Premium' in allData['VehicleTypeDetails']:
                    date_wise_count_sheet.cell(i, 4, allData['VehicleTypeDetails']['Premium'])
                else:
                    date_wise_count_sheet.cell(i, 4, 0)
                for zone_key,zone_count in allData['ZoneViseEntryDetails'].items():
                    for col_index,value in zone_header_dict.items():
                        if zone_key == value:
                            a5 = date_wise_count_sheet.cell(i, col_index, zone_count)
                i=i+1
            query = {'from':f_date, 'to':t_date,'site_id':site_id }
            headers={'auth_token': auth_token,'User-Agent':'Mozilla/5.0', 'Connection': 'keep-alive'}
            
            now = datetime.now()
            dt_string = now.strftime("%Y-%m-%d")
            vehicle_sheet.cell(3, 4, site_name)
            vehicle_sheet.cell(4, 4, dt_string)
            vehicle_sheet.cell(5, 4, f.strftime("%d-%m-%Y %H:%M:%S"))
            vehicle_sheet.cell(6, 4, t.strftime("%d-%m-%Y %H:%M:%S"))
            now_api = datetime.now()
            vehicle_result=elkVehicleAPI(site_id,f,t)
            now_api_res = datetime.now()
            TotalVehicleCount =0
            if len(vehicle_count_response['aggregations']['TotalVehicleCount']['buckets'])>0 :
                TotalVehicleCount=vehicle_count_response['aggregations']['TotalVehicleCount']['buckets'][0]["doc_count"]
            vehicle_sheet.cell(10, 3, TotalVehicleCount )
            VehicleTypeDetails = vehicle_count_response['aggregations']['VehicleType']['buckets']
            ZoneViseEntryDetails = vehicle_count_response['aggregations']['ZoneDetails']['buckets']
            
            vt_start_col=4
            for typeData in VehicleTypeDetails:
                vehicle_sheet.cell(9,vt_start_col, typeData["key"])
                vehicle_sheet.cell(10, vt_start_col, typeData["doc_count"])
                vt_start_col=vt_start_col+1
            zd_start_col=6
            for zoneData in ZoneViseEntryDetails:
                vehicle_sheet.cell(9, zd_start_col, zoneData["key"])
                vehicle_sheet.cell(10, zd_start_col, zoneData["doc_count"])
                zd_start_col=zd_start_col+1
            row=13
            if isimageRequired:
                cell_width = 100.0
                vehicle_sheet.column_dimensions['L'].width = cell_width/5.55
                vehicle_sheet.cell(12, 15, "VehicleImage")
            VehiclesANPRdata=[]
            for vehicleData in vehicle_result:
                vehicle_sheet.cell(row, 2, s_no)
                tempObj={}
                vehicle_sheet.cell(row, 6, vehicleData['Area'])
                if (vehicleData['Direction'] == 'IN') :
                    vehicle_sheet.cell(row, 4, vehicleData['Area'])
                    vehicle_sheet.cell(row, 6, "")
                    vehicle_sheet.cell(row, 3, verifyValue(vehicleData['CheckInTime']))
                else: 
                    vehicle_sheet.cell(row, 4, vehicleData['LaneIn'])
                    vehicle_sheet.cell(row, 6, vehicleData['Area'])
                    vehicle_sheet.cell(row, 3, verifyValue(vehicleData['VehicleCheckInTime']))
                if 'CheckOutTime' in vehicleData:
                    vehicle_sheet.cell(row, 5, vehicleData['CheckOutTime'])
                else:
                    vehicle_sheet.cell(row, 5, "")
                if (vehicleData['ANPR_Marge'] != None and vehicleData['ANPR_Marge'].startswith("None_") == True) :
                    vehicle_sheet.cell(row, 7, "N/A")
                    vehicle_sheet.cell(row, 11, 1)
                    vehicle_sheet.cell(row, 9, verifyValue(vehicleData['DateTime']))
                else :
                    vehicle_sheet.cell(row, 7, vehicleData['ANPR'])
                    #vehicle_sheet.cell(row, 9, vehicleData['FirstVisit'])
                    tempObj["ANPR"]=vehicleData['ANPR']
                    tempObj['DateTime']=vehicleData['DateTime']
                    oldRecords=[d for d in VehiclesANPRdata if d["ANPR"] == str(vehicleData['ANPR'])]
                    if (len(oldRecords) == 0):
                        vehicle_sheet.cell(row, 11, 1)
                        tempObj['RecentVisit'] = 1
                        vehicle_sheet.cell(row, 9, verifyValue(vehicleData['DateTime']))
                        
                    else :
                        odRecent=oldRecords[len(oldRecords) - 1]["RecentVisit"] + 1
                        vehicle_sheet.cell(row, 11, odRecent)
                        tempObj['RecentVisit'] = odRecent
                        vehicle_sheet.cell(row, 9, verifyValue(oldRecords[0]["DateTime"]))
                    VehiclesANPRdata.append(tempObj)
                if ('DwellTime' in vehicleData): 
                    vehicle_sheet.cell(row, 8, vehicleData['DwellTime'])
                else:
                    vehicle_sheet.cell(row, 8, "")
                vehicle_sheet.cell(row, 10, verifyValue(vehicleData['DateTime']))
                #vehicle_sheet.cell(row, 12, vehicleData['VisitCount'])
                if isimageRequired:
                    imageName = vehicleData['NumberPlateImage'].rsplit('/', 1)[-1]
                    path=VehicleImagePath+""+imageName
                    #path='/home/smag/Downloads/json/images.png'
                    try:    
                        with Image.open(str(path)) as img2:
                            width_100 = img2.width
                            height_100 = img2.height
                            width_100 = width_100*1.332
                            height_100 = height_100*1.332
                            cell_height = int(height_100 * (cell_width / width_100))
                            vehicle_sheet.row_dimensions[row].height = cell_height
                            x_scale = cell_width/width_100
                            y_scale = cell_height/height_100
                            
                            img = drawing.image.Image(path)
                            img.width = x_scale*width_100*1.33
                            img.height = y_scale*height_100*1.33
                            img.anchor = 'L'+str(row) # Or whatever cell location you want to use.
                            vehicle_sheet.add_image(img)
                    except Exception as ex:
                        vehicle_sheet.cell(row, 12, "Image Not Found")
                row=row+1
                s_no=s_no+1
        vehicle_sheet.cell(10, 10, onlinePayment)
        vehicle_sheet.cell(10, 11, cashPayment)
        vehicle_sheet.cell(10, 12, cashPayment+onlinePayment)
        site_name=site_name.replace(" ", "-") 
        ed_date= datetime. strptime(data["fileName"], '%Y-%m-%d') 
     
        report_wb.save('./report/Smarg_{}_Vehicles_Report_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
        time.sleep(2.4)
        print('Smarg_{}_Vehicles_Report_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
        print("Report Generate successfull")
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
    
