import json
import os
import sys
import requests
# Writing to an excel  
# sheet using Python 
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
PC_REQUEST_PATH = "./requests/footfall_Distribution_Request.json"
date=None
priorPeriodReportEnable=False
numberOfDayDelta=0
url = "http://localhost:9200/filebeat-*/_search/?size="
# url = "http://smargtech.in:36523/filebeat-*/_search/?size="
#url = "http://smargtechnologies.in:36523/filebeat-*/_search/?size="
#url = "http://27.107.152.220:36523/filebeat-*/_search/?size="
headers = {'content_type': 'application/json','Authorization':'Basic ZWxhc3RpYzpzbWFyZzEyM3NtYXJnIw=='}
def create_request_for_all_data(site_id, zone_ids, from_date, to_date):
    #json=json.loads("{\"query\":{\"bool\":{\"must\":[],\"filter\": [{\"term\": {\"SiteId\": 111}},{\"term\": {\"ZoneId\": 1}},{\"term"\: {\"isViolation\": 0}},{\"term\": {\"Direction.keyword\": \"IN\"}},{\"term\": {\"NotificationType\": \"7\"}},{\"range\": {\"DateTime\": {\"gte\": \"2020-09-26T00:00:00Z\",\"lte\": \"2020-12-26T11:00:00Z\",}}},],}},\"sort\": [{\"@timestamp\": {\"order\": \"desc\"}}],}")
    json_obj=None
    with open(PC_REQUEST_PATH) as json_data:
        json_obj = json.load(json_data)
    json_obj["query"]["bool"]["filter"][0]["terms"]["SiteId"]=[site_id]
    json_obj["query"]["bool"]["filter"][1]["terms"]["CameraID"]=zone_ids
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["gte"]="{:02d}-{:02d}-{:02d}T".format(from_date.year, from_date.month, from_date.day)+"{:02d}:{:02d}:{:02d}.000Z".format(from_date.hour, from_date.minute,from_date.second)
    json_obj["query"]["bool"]["filter"][5]["range"]["DateTime"]["lte"]="{:02d}-{:02d}-{:02d}T".format(to_date.year, to_date.month, to_date.day)+"{:02d}:{:02d}:{:02d}.999Z".format(to_date.hour, to_date.minute,to_date.second)
    #print("json_obj",json_obj)
    return json_obj
def mothly_weeklyData(site_id, f_date, t_date):
    w_row=2
    # weekCount = report_wb['WeekCount']
    weekCount =None
    if priorPeriodReportEnable:
        weekCount = report_wb['PriorCount']
    else:
        weekCount = report_wb['Week_Days_Wise_Passage_Count']
    fromTimeOfSite=f_date.time()
    toTimeOfSite=t_date.time()
    fromDateOfSite=f_date.date()
    toDateOfSite=t_date.date()
    dayTime=toTimeOfSite.hour-fromTimeOfSite.hour
    isSameDay=False
    if(dayTime>0):
        isSameDay=True
    delta = timedelta(days=1)
    DLC=''
    DLW=''
    DLM=''
    while fromDateOfSite <= toDateOfSite:
        if not isSameDay and fromDateOfSite==toDateOfSite:
            fromDateOfSite +=delta
        else:
            dateWiseFromDate=str(fromDateOfSite)+"T"+str(fromTimeOfSite)
            dateWiseToDate=""
            if isSameDay:
                dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
                fromDateOfSite +=delta
            else:
                fromDateOfSite +=delta
                dateWiseToDate=str(fromDateOfSite)+"T"+str(toTimeOfSite)
            #Current Date Count
            dateWiseFromDate =datetime.strptime(dateWiseFromDate, "%Y-%m-%dT%H:%M:%S")
            dateWiseToDate =datetime.strptime(dateWiseToDate, "%Y-%m-%dT%H:%M:%S")
            current_payload = create_request_for_all_data(site_id, zonesId, dateWiseFromDate, dateWiseToDate)
            current_response = requests.get(url+"0",json=current_payload,headers=headers,timeout=(300,300))
            current_response=current_response.json()
            open("./responsejsndata.json","w").write(json.dumps(current_response))
            current_total=current_response["aggregations"]["total_count"]["value"]
            
            if priorPeriodReportEnable:
                #Past Week Date Count
                datePastWeekFromDate = dateWiseFromDate - timedelta(days=numberOfDayDelta)
                datePastWeekToDate = dateWiseToDate - timedelta(days=numberOfDayDelta)
                past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
                past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300))
                past_week_response=past_week_response.json()
                past_week_total=past_week_response["aggregations"]["total_count"]["value"]
                
                new_f_date = datetime.strftime(dateWiseFromDate, "%d-%m-%Y")
                weekCount.cell(w_row+1, 1, w_row-1)
                weekCount.cell(w_row+1, 2, new_f_date)
                weekCount.cell(w_row+1, 3, current_total)
                weekCount.cell(w_row+1, 4, past_week_total)
                if w_row==2:
                    DLC=dateWiseFromDate.strftime('%Y-%m-%d')
                    DLW=datePastWeekFromDate.strftime('%Y-%m-%d')
                if True:
                    DLC_1="Selected ( "+str(DLC)+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
                    DLW_1="Prior Period ( "+str(DLW)+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
                    weekCount.cell(1, 3,DLC_1 )
                    weekCount.cell(1, 4,DLW_1 )
            else:
                #Past Week Date Count
                datePastWeekFromDate = dateWiseFromDate - timedelta(days=7)
                datePastWeekToDate = dateWiseToDate - timedelta(days=7)
                past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
                past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300))
                past_week_response=past_week_response.json()
                past_week_total=past_week_response["aggregations"]["total_count"]["value"]
                
                #Past Month Count
                # datePastMonthFromDate = dateWiseFromDate + relativedelta(months=-1)
                # datePastMonthToDate = dateWiseToDate + relativedelta(months=-1)
                datePastMonthFromDate = dateWiseFromDate - timedelta(days=28)
                datePastMonthToDate = dateWiseToDate - timedelta(days=28)
                past_month_payload = create_request_for_all_data(site_id, zonesId, datePastMonthFromDate, datePastMonthToDate)
                past_month_response = requests.get(url+"0",json=past_month_payload,headers=headers,timeout=(300,300))
                past_month_response=past_month_response.json()
                past_month_total=past_month_response["aggregations"]["total_count"]["value"]
                
                new_f_date = datetime.strftime(dateWiseFromDate, "%d-%m-%Y")
                weekCount.cell(w_row+1, 1, w_row-1)
                weekCount.cell(w_row+1, 2, dateWiseFromDate.strftime('%A'))
                weekCount.cell(w_row+1, 3, current_total)
                weekCount.cell(w_row+1, 4, past_week_total)
                weekCount.cell(w_row+1, 5, past_month_total)
                if w_row==2:
                    DLC=dateWiseFromDate.strftime('%Y-%m-%d')
                    DLW=datePastWeekFromDate.strftime('%Y-%m-%d')
                    DLM=datePastMonthFromDate.strftime('%Y-%m-%d')
                    
                if True:
                    DLC_1="Selected ( "+str(DLC)+" TO "+str(dateWiseToDate.strftime('%Y-%m-%d'))+" )"
                    DLW_1="Last Week ( "+str(DLW)+" TO "+str(datePastWeekToDate.strftime('%Y-%m-%d'))+" )"
                    DLM_1="Last Month ( "+str(DLM)+" TO "+str(datePastMonthToDate.strftime('%Y-%m-%d'))+" )"
                    weekCount.cell(1, 3,DLC_1 )
                    weekCount.cell(1, 4,DLW_1 )
                    weekCount.cell(1, 5,DLM_1 )
            
            # allDataCountDict[new_f_date] = allCountsDict 
        w_row=w_row+1
     
def get_percentage_diff(previous, current):
    try:
        percentage = (previous-current)/current * 100
    except ZeroDivisionError:
        percentage = float(100)    
    return round(percentage,2)
def zone_wise_details(site_id, f_date, t_date):
    # t3 = t_date - f_date
    # t3 = t3.days
    datePastWeekFromDate=''
    datePastWeekToDate=''
    datePastYearFromDate=''
    datePastYearToDate=''
    fromTimeOfSite=f_date.time()
    toTimeOfSite=t_date.time()
    fromDateOfSite=f_date.date()
    toDateOfSite=t_date.date()
    dateWiseFromDate=str(fromDateOfSite)+"T"+str(fromTimeOfSite)
    dateWiseToDate=str(toDateOfSite)+"T"+str(toTimeOfSite)       
    #Current Date Count
    dateWiseFromDate =datetime.strptime(dateWiseFromDate, "%Y-%m-%dT%H:%M:%S")
    dateWiseToDate =datetime.strptime(dateWiseToDate, "%Y-%m-%dT%H:%M:%S")
    current_payload = create_request_for_all_data(site_id, zonesId, dateWiseFromDate, dateWiseToDate)
    current_response = requests.get(url+"0",json=current_payload,headers=headers,timeout=(300,300))
    current_response=current_response.json()
    t_res=current_response["hits"]["total"]["value"]
    if t_res is None or t_res == 0:
        return None
    allZoneCountCDate=current_response["aggregations"]["total_count"]["value"]
    allZoneCountCurrentDate=current_response["aggregations"]["store_wise_total"]["buckets"]
    if priorPeriodReportEnable:
        #Past Week Date Count
        datePastWeekFromDate = dateWiseFromDate - timedelta(days=numberOfDayDelta)
        datePastWeekToDate = dateWiseToDate - timedelta(days=numberOfDayDelta)
        past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
        past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300))
        past_week_response=past_week_response.json()
        allZoneCountPastWeek=past_week_response["aggregations"]["store_wise_total"]["buckets"]
        allZonePastWCount=past_week_response["aggregations"]["total_count"]["value"]
    else:
        #Past Week Date Count
        datePastWeekFromDate = dateWiseFromDate - timedelta(days=7)
        datePastWeekToDate = dateWiseToDate - timedelta(days=7)
        past_week_payload = create_request_for_all_data(site_id, zonesId, datePastWeekFromDate, datePastWeekToDate)
        past_week_response = requests.get(url+"0",json=past_week_payload,headers=headers,timeout=(300,300))
        past_week_response=past_week_response.json()
        allZoneCountPastWeek=past_week_response["aggregations"]["store_wise_total"]["buckets"]
        allZonePastWCount=past_week_response["aggregations"]["total_count"]["value"]
        #Past Year Count
        datePastYearFromDate = dateWiseFromDate - relativedelta(years=1)
        datePastYearToDate = dateWiseToDate - relativedelta(years=1)
        past_year_payload = create_request_for_all_data(site_id, zonesId, datePastYearFromDate, datePastYearToDate)
        past_year_response = requests.get(url+"0",json=past_year_payload,headers=headers,timeout=(300,300))
        past_year_response=past_year_response.json()
        allZoneCountPastYear=past_year_response["aggregations"]["store_wise_total"]["buckets"]
        allZonePastYCount=past_year_response["aggregations"]["total_count"]["value"]
if __name__=='__main__':
    if sys.argv[1] is not None:
        jsondata=json.loads(str(sys.argv[1]))
        # data={"site_id":112, "site_name":"Inorbit Mall Hyderabad","from_date": "2024-04-01 00:00:01","to_date": "2024-04-09 12:00:00","zones":[]}
        site_id = jsondata["site_id"]
        site_name = jsondata["site_name"]
        f_date= datetime. strptime(jsondata["from_date"], '%Y-%m-%d %H:%M:%S')
        t_date = datetime.strptime(jsondata["to_date"], '%Y-%m-%d %H:%M:%S')
        f = f_date
        t=t_date
        zones = jsondata["zones"]
        dayDiffernce = t - f
        numberOfDayDelta=dayDiffernce.days
        fromTimeOfSite=f.time()
        toTimeOfSite=t.time()
        dayTime=toTimeOfSite.hour-fromTimeOfSite.hour
        t2=t
        if (dayTime==0):
            t2 = t - timedelta(days=1)
        if(dayTime>=0):
           pass
        else: 
            numberOfDayDelta=numberOfDayDelta+1
            t2 = t - timedelta(days=1)
        if(numberOfDayDelta>7):
            priorPeriodReportEnable=False
        try:
            # Create  copy of template workbook required for generating report
            report_wb = load_workbook("./requests/footfall_Distribution.xlsx")
            # ws1 = report_wb.create_sheet("Mysheet")
            if priorPeriodReportEnable:
                report_wb = load_workbook("./requests/footfall_Distribution_Prior_Period.xlsx")
            #report_wb = copy(template_wb)
        except Exception as ex:
            print("Exception while opening template excel {}".format(ex))
 
        if f_date is not None:
            #Open Footfall dashboard sheet and update the site title and dates
            #Inorbit Mall (Vadodara) - Footfall Analytics Report 
            #(for the period 1 Jan'21 - 31 Jan'21)
            # footfall_dash = report_wb['Summary']
            footfall_dash = report_wb['Passage_Summary']
            footfall_dash.cell(3, 3, "{} - Footfall Distribution Report \n (for the period {} - {})".format(site_name,f.strftime("%d %b'%y"),t2.strftime("%d %b'%y")))
            s_no = 1
            row=2
            
            '''while f_date <= t_date:
                print(f_date)
                print(t_date)
                f_date = f_date + timedelta(days=1)'''
            
            #while f_date <= t_date:
            #for i in range(f_date.day, t_date.day+1):            
                # open Footfall Data sheet. 
            # footfall_sheet = report_wb['Store_Footfall_Data']
            from_date=f_date+timedelta(hours=00)
            to_date=f_date+timedelta(hours=23)
            to_date=to_date+timedelta(minutes=59)
            to_date=to_date+timedelta(seconds=59)
            zonesId = [o["id"] for o in zones]
            #Edited by Sandeep
            mothly_weeklyData(site_id, f_date, t_date)
           
            #Edited by Shubham
            zone_wise_details(site_id, f_date, t_date)
            #Edited by Mukesh
       
            font = Font( bold=True)
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            footfall_payload = create_request_for_all_data(site_id, zonesId, f_date, t_date)
            response = requests.get(url+"0",json=footfall_payload,headers=headers,timeout=(300,300))
            footfall_result=response.json()
            #print("========== footfall_result ======",footfall_result)
            # floor_data=footfall_result["aggregations"]["floor_wise_total"]["buckets"]
            # # floor_wise_sheet = report_wb.create_sheet("Block_Floor_Wise_Footfall_Data")
            
            dateInterval=footfall_result["aggregations"]["day_wise_passage_and_store_footfall_count"]["buckets"]
            for dateWisetData in dateInterval:
                if dateWisetData['passage_filter']['doc_count']>0:
                    allZoneCountDetails=dateWisetData['passage_filter']["PassageDayWiseCount"]["buckets"]
                    for zoneCountData in allZoneCountDetails:
                        footfall_sheet = report_wb['Passages_Count_Day_Wise']
                        footfall_sheet.cell(row, 1, s_no)
                        footfall_sheet.cell(row, 2, dateWisetData["key_as_string"].split("T")[0])
                        footfall_sheet.cell(row, 3, zoneCountData['key'])
                        footfall_sheet.cell(row, 4, zoneCountData["doc_count"])
                        genderData=zoneCountData['PassageGender']['buckets']
                        for gender in genderData:
                            if gender["key"] == "Male":
                                footfall_sheet.cell(row, 5, gender["doc_count"])
                            if gender["key"] == "Female":
                                footfall_sheet.cell(row, 6, gender["doc_count"])
                        
                        ageGroupData=zoneCountData['PassageAgeGroup']['buckets']
                        for ageData in ageGroupData:
                            if ageData["key"] == "0-15":
                                footfall_sheet.cell(row, 7, ageData["doc_count"])
                            if ageData["key"] == "16-35":
                                footfall_sheet.cell(row, 8, ageData["doc_count"])
                            if ageData["key"] == "16-25":
                                footfall_sheet.cell(row, 9, ageData["doc_count"])
                            if ageData["key"] == "26-35":
                                footfall_sheet.cell(row, 10, ageData["doc_count"])
                            if ageData["key"] == "36-55":
                                footfall_sheet.cell(row, 11, ageData["doc_count"])
                            if ageData["key"] == "55+":
                                footfall_sheet.cell(row, 12, ageData["doc_count"])
                        row=row+1
                        s_no=s_no+1
                
                if dateWisetData['store_filter']['doc_count']>0:
                    footfall_sheet = report_wb['Store_Count_Day_Wise']
                    allZoneCountDetails=dateWisetData['store_filter']["StoreDayWiseCount"]["buckets"]
                    for zoneCountData in allZoneCountDetails:
                        footfall_sheet.cell(row, 1, s_no)
                        footfall_sheet.cell(row, 2, dateWisetData["key_as_string"].split("T")[0])
                        footfall_sheet.cell(row, 3, zoneCountData['key'])
                        footfall_sheet.cell(row, 4, zoneCountData["doc_count"])
                        genderData=zoneCountData['StoreGender']['buckets']
                        for gender in genderData:
                            if gender["key"] == "Male":
                                footfall_sheet.cell(row, 5, gender["doc_count"])
                            if gender["key"] == "Female":
                                footfall_sheet.cell(row, 6, gender["doc_count"])
                        
                        ageGroupData=zoneCountData['StoreAgeGroup']['buckets']
                        for ageData in ageGroupData:
                            if ageData["key"] == "0-15":
                                footfall_sheet.cell(row, 7, ageData["doc_count"])
                            if ageData["key"] == "16-35":
                                footfall_sheet.cell(row, 8, ageData["doc_count"])
                            if ageData["key"] == "16-25":
                                footfall_sheet.cell(row, 9, ageData["doc_count"])
                            if ageData["key"] == "26-35":
                                footfall_sheet.cell(row, 10, ageData["doc_count"])
                            if ageData["key"] == "36-55":
                                footfall_sheet.cell(row, 11, ageData["doc_count"])
                            if ageData["key"] == "55+":
                                footfall_sheet.cell(row, 12, ageData["doc_count"])
                        row=row+1
                        s_no=s_no+1
                
                f_date = f_date + timedelta(days=1)
            # footfall_Hourly_sheet = report_wb['Store_Footfall_Hourly_Data']
            allStoreZoneCountCurrentDate=footfall_result["aggregations"]["store_wise_total"]["buckets"]
            locationWiseAgeGenderDic={}
            locationCol=2
            if len(allStoreZoneCountCurrentDate)>0:
                for zoneCountData in allStoreZoneCountCurrentDate:
                    ws1 = report_wb.create_sheet(zoneCountData["key"])
                    # footfall_Hourly_sheet.cell(2, locationCol, zoneCountData["key"])
                    ws1.cell(1, 2, 'Location Name').font=font
                    ws1.cell(1, 3, zoneCountData["key"]).font=font
                    ws1.cell(3, locationCol, 'From').border=border
                    ws1.cell(3, locationCol+1, 'To').border=border
                    ws1.cell(3, locationCol+2, 'Total').border=border
                    ws1.cell(3, locationCol+3, 'Male').border=border
                    ws1.cell(3, locationCol+4, 'Female').border=border
                    ws1.cell(3, locationCol+5, 'Age Group (0-15)').border=border
                    ws1.cell(3, locationCol+6, 'Age Group (16-25)').border=border
                    ws1.cell(3, locationCol+7, 'Age Group (26-35)').border=border
                    ws1.cell(3, locationCol+8, 'Age Group (36-55)').border=border
                    ws1.cell(3, locationCol+9, 'Age Group (55+)').border=border
                    ws1.column_dimensions[chr(65 + 1)].width = 25
                    ws1.column_dimensions[chr(65 + 2)].width = 25  
                    ws1.column_dimensions[chr(65 + 6)].width = 18
                    ws1.column_dimensions[chr(65 + 7)].width = 18  
                    ws1.column_dimensions[chr(65 + 8)].width = 18
                    ws1.column_dimensions[chr(65 + 9)].width = 18 
                    ws1.column_dimensions[chr(65 + 10)].width = 18  
                    # row_dimensions
                    ws1['B3'].font=font
                    ws1['C3'].font=font
                    ws1['D3'].font=font
                    ws1['E3'].font=font
                    ws1['F3'].font=font
                    ws1['G3'].font=font
                    ws1['H3'].font=font
                    ws1['I3'].font=font
                    ws1['J3'].font=font
                    ws1['K3'].font=font
                    # print("--------",zoneCountData["key"])
                    locationWiseAgeGenderDic[zoneCountData["key"]]=locationCol
                    locationCol=locationCol
                    hr_col=4
                    
                    hr_row=4
                    time_row_dic={}
                    new_hours_data=f
                    while new_hours_data < t:
                        # print("step ---1",new_hours_data)
                        time_row_dic[new_hours_data]=hr_row
                        ws1.cell(hr_row, 2, new_hours_data).border=border
                        # print("step ---2",new_hours_data)
                        new_hours_data = new_hours_data+ timedelta(hours=1)
                        ws1.cell(hr_row, 3, new_hours_data).border=border
                        # print("step ---3",new_hours_data)
                    
                        hr_row=hr_row + 1
                    
                    for i in range(4,hr_row):
                        for j in range(hr_col,hr_col+8):
                            ws1.cell(i, j, 0).border=border
            allPassageZoneCountCurrentDate=footfall_result["aggregations"]["passage_wise_total"]["buckets"]
            locationWiseAgeGenderDic={}
            locationCol=2
            if len(allPassageZoneCountCurrentDate)>0:
                for zoneCountData in allPassageZoneCountCurrentDate:
                    ws1 = report_wb.create_sheet(zoneCountData["key"])
                    # footfall_Hourly_sheet.cell(2, locationCol, zoneCountData["key"])
                    ws1.cell(1, 2, 'Location Name').font=font
                    ws1.cell(1, 3, zoneCountData["key"]).font=font
                    ws1.cell(3, locationCol, 'From').border=border
                    ws1.cell(3, locationCol+1, 'To').border=border
                    ws1.cell(3, locationCol+2, 'Total').border=border
                    ws1.cell(3, locationCol+3, 'Male').border=border
                    ws1.cell(3, locationCol+4, 'Female').border=border
                    ws1.cell(3, locationCol+5, 'Age Group (0-15)').border=border
                    ws1.cell(3, locationCol+6, 'Age Group (16-25)').border=border
                    ws1.cell(3, locationCol+7, 'Age Group (26-35)').border=border
                    ws1.cell(3, locationCol+8, 'Age Group (36-55)').border=border
                    ws1.cell(3, locationCol+9, 'Age Group (55+)').border=border
                    ws1.column_dimensions[chr(65 + 1)].width = 25
                    ws1.column_dimensions[chr(65 + 2)].width = 25  
                    ws1.column_dimensions[chr(65 + 6)].width = 18
                    ws1.column_dimensions[chr(65 + 7)].width = 18  
                    ws1.column_dimensions[chr(65 + 8)].width = 18
                    ws1.column_dimensions[chr(65 + 9)].width = 18 
                    ws1.column_dimensions[chr(65 + 10)].width = 18  
                    # row_dimensions
                    ws1['B3'].font=font
                    ws1['C3'].font=font
                    ws1['D3'].font=font
                    ws1['E3'].font=font
                    ws1['F3'].font=font
                    ws1['G3'].font=font
                    ws1['H3'].font=font
                    ws1['I3'].font=font
                    ws1['J3'].font=font
                    ws1['K3'].font=font
                    locationWiseAgeGenderDic[zoneCountData["key"]]=locationCol
                    locationCol=locationCol
                    hr_col=4
                    
                    hr_row=4
                    time_row_dic={}
                    new_hours_data=f
                    while new_hours_data < t:
                        time_row_dic[new_hours_data]=hr_row
                        ws1.cell(hr_row, 2, new_hours_data).border=border
                        new_hours_data = new_hours_data+ timedelta(hours=1)
                        ws1.cell(hr_row, 3, new_hours_data).border=border
                        
                        hr_row=hr_row + 1
                    
                    for i in range(4,hr_row):
                        for j in range(hr_col,hr_col+8):
                            ws1.cell(i, j, 0).border=border
            HoulyFootfallCountDetails=footfall_result["aggregations"]["hour_wise_passage_and_store_footfall_count"]["buckets"]
            for each_hour_data in HoulyFootfallCountDetails:
                tm=datetime.strptime(each_hour_data["key_as_string"], '%Y-%m-%d %H:%M:%S')
                if (tm in time_row_dic):
                    # ZoneDetails_Hourly_Count=footfall_result["aggregations"]["ZoneDetails_Hourly_Count"]["buckets"]
                    storeZoneDetails_Hourly_Count=each_hour_data["store_filter"]["StoreHourWiseCount"]["buckets"]
                    if len(storeZoneDetails_Hourly_Count)>0:
                        for ZonesData in storeZoneDetails_Hourly_Count:
                            # zoneColNo=locationWiseAgeGenderDic[ZonesData['key']]
                            zoneColNo=4
                            # footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, ZonesData['key'])
                            footfall_Hourly_sheet = report_wb[ZonesData['key']]
                            footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo , ZonesData["doc_count"])
                            GenderHourly=ZonesData['GenderHourly']['buckets']
                            for gender in GenderHourly:
                                if gender["key"] == "Male":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+1, gender["doc_count"])
                                if gender["key"] == "Female":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+2, gender["doc_count"])
                            AgeGroupHourly=ZonesData['AgeGroupHourly']['buckets']
                            for ageData in AgeGroupHourly:
                                if ageData["key"] == "0-15":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+3, ageData["doc_count"])
                                if ageData["key"] == "16-25":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+4, ageData["doc_count"])
                                if ageData["key"] == "26-35":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+5, ageData["doc_count"])
                                if ageData["key"] == "36-55":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+6, ageData["doc_count"])
                                if ageData["key"] == "55+":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+7, ageData["doc_count"])
                      # ZoneDetails_Hourly_Count=footfall_result["aggregations"]["ZoneDetails_Hourly_Count"]["buckets"]
                    passageZoneDetails_Hourly_Count=each_hour_data["passage_filter"]["PassageHourWiseCount"]["buckets"]
                    if len(passageZoneDetails_Hourly_Count)>0:
                        for ZonesData in passageZoneDetails_Hourly_Count:
                            # zoneColNo=locationWiseAgeGenderDic[ZonesData['key']]
                            zoneColNo=4
                            # footfall_Hourly_sheet.cell(time_row_dic[tm], hr_col, ZonesData['key'])
                            footfall_Hourly_sheet = report_wb[ZonesData['key']]
                            footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo , ZonesData["doc_count"])
                            GenderHourly=ZonesData['GenderHourly']['buckets']
                            for gender in GenderHourly:
                                if gender["key"] == "Male":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+1, gender["doc_count"])
                                if gender["key"] == "Female":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+2, gender["doc_count"])
                            AgeGroupHourly=ZonesData['AgeGroupHourly']['buckets']
                            for ageData in AgeGroupHourly:
                                if ageData["key"] == "0-15":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+3, ageData["doc_count"])
                                if ageData["key"] == "16-25":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+4, ageData["doc_count"])
                                if ageData["key"] == "26-35":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+5, ageData["doc_count"])
                                if ageData["key"] == "36-55":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+6, ageData["doc_count"])
                                if ageData["key"] == "55+":
                                    footfall_Hourly_sheet.cell(time_row_dic[tm], zoneColNo+7, ageData["doc_count"])
            
            total_footfall_sheet=report_wb['Total_Footfall_Distribution']
            rowno=5
            floor_wise_data=footfall_result["aggregations"]['floor_wise_total']['buckets']
            store_wise_data=footfall_result["aggregations"]['store_wise_total']['buckets']
            block_wise_data=footfall_result["aggregations"]['block_wise_total']['buckets']
            passage_wise_data=footfall_result["aggregations"]['passage_wise_total']['buckets']
            if len(floor_wise_data)>0:
                rowno=5
                sNo=1
                colNo=11
                for flrData in floor_wise_data:
                    total_footfall_sheet.cell(rowno,colNo+1,sNo)
                    total_footfall_sheet.cell(rowno,colNo+2,flrData['key'])
                    total_footfall_sheet.cell(rowno,colNo+3,flrData['doc_count'])
                    genderFloorData=flrData['GenderHourly']['buckets']
                    for data in genderFloorData:
                        if data['key']=="Male":
                            total_footfall_sheet.cell(rowno,colNo+4,data['doc_count'])
                        if data['key']=="Female":
                            total_footfall_sheet.cell(rowno,colNo+5,data['doc_count'])
                    AgeGroupHourly=flrData['AgeGroupHourly']['buckets']
                    for ageData in AgeGroupHourly:
                        if ageData["key"] == "0-15":
                            total_footfall_sheet.cell(rowno,colNo+ 6, ageData["doc_count"])
                        if ageData["key"] == "16-25":
                            total_footfall_sheet.cell(rowno,colNo+ 7, ageData["doc_count"])
                        if ageData["key"] == "26-35":
                            total_footfall_sheet.cell(rowno,colNo+ 8, ageData["doc_count"])
                        if ageData["key"] == "36-55":
                            total_footfall_sheet.cell(rowno,colNo+ 9, ageData["doc_count"])
                        if ageData["key"] == "55+":
                            total_footfall_sheet.cell(rowno,colNo+ 10, ageData["doc_count"])
                
                    sNo+=1
                    rowno+=1
            if len(store_wise_data)>0:
                rowno=5
                sNo=1
                colNo=33
                for storData in store_wise_data:
                    total_footfall_sheet.cell(rowno,colNo+1,sNo)
                    total_footfall_sheet.cell(rowno,colNo+2,storData['key'])
                    total_footfall_sheet.cell(rowno,colNo+3,storData['doc_count'])
                    genderFloorData=storData['GenderHourly']['buckets']
                    for data in genderFloorData:
                        if data['key']=="Male":
                            total_footfall_sheet.cell(rowno,colNo+4,data['doc_count'])
                        if data['key']=="Female":
                            total_footfall_sheet.cell(rowno,colNo+5,data['doc_count'])
                    AgeGroupHourly=storData['AgeGroupHourly']['buckets']
                    for ageData in AgeGroupHourly:
                        if ageData["key"] == "0-15":
                            total_footfall_sheet.cell(rowno,colNo+ 6, ageData["doc_count"])
                        if ageData["key"] == "16-25":
                            total_footfall_sheet.cell(rowno,colNo+ 7, ageData["doc_count"])
                        if ageData["key"] == "26-35":
                            total_footfall_sheet.cell(rowno,colNo+ 8, ageData["doc_count"])
                        if ageData["key"] == "36-55":
                            total_footfall_sheet.cell(rowno,colNo+ 9, ageData["doc_count"])
                        if ageData["key"] == "55+":
                            total_footfall_sheet.cell(rowno,colNo+ 10, ageData["doc_count"])
                
                    sNo+=1
                    rowno+=1
            if len(block_wise_data)>0:
                rowno=5
                sNo=1
                colNo=0
                for blockData in block_wise_data:
                    total_footfall_sheet.cell(rowno,colNo+1,sNo)
                    total_footfall_sheet.cell(rowno,colNo+2,blockData['key'])
                    total_footfall_sheet.cell(rowno,colNo+3,blockData['doc_count'])
                    genderFloorData=blockData['GenderHourly']['buckets']
                    for data in genderFloorData:
                        if data['key']=="Male":
                            total_footfall_sheet.cell(rowno,colNo+4,data['doc_count'])
                        if data['key']=="Female":
                            total_footfall_sheet.cell(rowno,colNo+5,data['doc_count'])
                    AgeGroupHourly=blockData['AgeGroupHourly']['buckets']
                    for ageData in AgeGroupHourly:
                        if ageData["key"] == "0-15":
                            total_footfall_sheet.cell(rowno,colNo+ 6, ageData["doc_count"])
                        if ageData["key"] == "16-25":
                            total_footfall_sheet.cell(rowno,colNo+ 7, ageData["doc_count"])
                        if ageData["key"] == "26-35":
                            total_footfall_sheet.cell(rowno,colNo+ 8, ageData["doc_count"])
                        if ageData["key"] == "36-55":
                            total_footfall_sheet.cell(rowno,colNo+ 9, ageData["doc_count"])
                        if ageData["key"] == "55+":
                            total_footfall_sheet.cell(rowno,colNo+ 10, ageData["doc_count"])
                
                    sNo+=1
                    rowno+=1
            
            
            if len(passage_wise_data)>0:
                rowno=5
                sNo=1
                colNo=22
                for pssgData in passage_wise_data:
                    total_footfall_sheet.cell(rowno,colNo+1,sNo)
                    total_footfall_sheet.cell(rowno,colNo+2,pssgData['key'])
                    total_footfall_sheet.cell(rowno,colNo+3,pssgData['doc_count'])
                    genderFloorData=pssgData['GenderHourly']['buckets']
                    for data in genderFloorData:
                        if data['key']=="Male":
                            total_footfall_sheet.cell(rowno,colNo+4,data['doc_count'])
                        if data['key']=="Female":
                            total_footfall_sheet.cell(rowno,colNo+5,data['doc_count'])
                    AgeGroupHourly=pssgData['AgeGroupHourly']['buckets']
                    for ageData in AgeGroupHourly:
                        if ageData["key"] == "0-15":
                            total_footfall_sheet.cell(rowno,colNo+ 6, ageData["doc_count"])
                        if ageData["key"] == "16-25":
                            total_footfall_sheet.cell(rowno,colNo+ 7, ageData["doc_count"])
                        if ageData["key"] == "26-35":
                            total_footfall_sheet.cell(rowno,colNo+ 8, ageData["doc_count"])
                        if ageData["key"] == "36-55":
                            total_footfall_sheet.cell(rowno,colNo+ 9, ageData["doc_count"])
                        if ageData["key"] == "55+":
                            total_footfall_sheet.cell(rowno,colNo+ 10, ageData["doc_count"])
                
                    sNo+=1
                    rowno+=1
                
            # for each_block_data in BlockWiseCategoryDetails:
            #     CategoryWiseDetails=each_block_data["Category"]["buckets"]
            #     floor_wise_sheet.cell(floorRow,7,srno).border=border
            #     floor_wise_sheet.cell(floorRow,8,each_block_data["key"]).border=border
            #     floor_wise_sheet.cell(floorRow,9, each_block_data["doc_count"]).border=border
            #     srno+1
            #     BlockCategorySheet.cell(block_row, 1, each_block_data["key"])
            #     BlockCategorySheet.cell(block_row, 2, each_block_data["doc_count"])
            #     for each_category_data in CategoryWiseDetails:
            #         if each_category_data["key"] not in cate_dic:
            #             #Category wise Block Traction
            #             OtherDicCat_Block[each_category_data["key"]]={}
                        
            #             #Block wise Category Traction
            #             cate_dic[each_category_data["key"]]=catcolNumber
            #             BlockCategorySheet.cell(2, catcolNumber, each_category_data["key"])
            #             catcolNumber=catcolNumber+1
            #         #Category wise Block Traction
            #         OtherDicCat_Block[each_category_data["key"]][each_block_data["key"]]=each_category_data["doc_count"]
            #         percentValue=int(each_category_data["doc_count"])/int(each_block_data["doc_count"])*100
            #         percentValue=round(percentValue, 2)
            #         percentValue=str(percentValue)+" %"
            #         BlockCategorySheet.cell(block_row, cate_dic[each_category_data["key"]], percentValue)
            #         #BlockCategorySheet.cell(block_row, cate_dic[each_category_data["key"]], each_category_data["doc_count"])
            #     block_row=block_row+1
            
            # #print("--",OtherDicCat_Block)
            # cat_rows=3
            # block_col_num=9
            # block_dic={}
            # for cat_name, block_data_details in OtherDicCat_Block.items():
            #     BlockCategorySheet.cell(cat_rows, 8, cat_name)
            #     for block_name, count in block_data_details.items():
            #         if block_name not in block_dic:
            #             block_dic[block_name]=block_col_num
            #             BlockCategorySheet.cell(2, block_col_num, block_name)
            #             block_col_num=block_col_num+1
                    
            #         BlockCategorySheet.cell(cat_rows,block_dic[block_name] , count)
            #     cat_rows=cat_rows+1
                   
        site_name=site_name.replace(" ", "-")  
        ed_date= datetime. strptime(jsondata["fileName"], '%Y-%m-%d')   
        #ed_date.strftime("%d-%m-%Y")    
        reportfile='./report/Smarg_{}_Footfall_Distribution_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y"))
        report_wb.save(reportfile)
        print('Smarg_{}_Footfall_Distribution_Reports_{}_{}.xls'.format(site_name,f.strftime("%d-%m-%Y"),ed_date.strftime("%d-%m-%Y")))
    else:
        print("Please pass the site id,site_name, zones, from date(YYYY-mm-dd HH:MM:ss) and to date(YYYY-mm-dd HH:MM:ss) for which reports needed.")
    
        
